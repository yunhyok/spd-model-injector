from __future__ import annotations

import csv
from dataclasses import dataclass, replace
import html
import io
from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QAction, QColor, QFont, QFontDatabase, QKeySequence, QPalette, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QCompleter,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QHeaderView,
    QInputDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QSplitter,
    QStatusBar,
    QTableWidget,
    QTableWidgetItem,
    QToolBar,
    QVBoxLayout,
    QWidget,
)

from spd_model_injector import __version__
from spd_model_injector.core.refdes_export import export_refdes_xlsx
from spd_model_injector.core.spd import PartialCktBlock, RefDesRecord, SpdInventory, read_block_body
from spd_model_injector.core.spice import ModelValidationError, prepare_model_for_partialckt
from spd_model_injector.ui.workers import ExportWorker, ScanWorker


class DropModelEditor(QPlainTextEdit):
    fileDropped = Signal(str)
    fileDropError = Signal(str)

    def __init__(self) -> None:
        super().__init__()
        self.setAcceptDrops(True)
        self.setPlaceholderText("Paste SPICE model text or drop a .mod/.txt file here.")
        font = QFontDatabase.systemFont(QFontDatabase.SystemFont.FixedFont)
        if not font.fixedPitch():
            font = QFont("Courier New")
        font.setStyleHint(QFont.StyleHint.Monospace)
        font.setFixedPitch(True)
        self.setFont(font)
        self.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)

    def dragEnterEvent(self, event) -> None:  # noqa: N802
        mime = event.mimeData()
        if mime.hasUrls():
            if any(url.isLocalFile() and Path(url.toLocalFile()).is_file() for url in mime.urls()):
                event.acceptProposedAction()
                return
            event.ignore()
            return
        if mime.hasText():
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dropEvent(self, event) -> None:  # noqa: N802
        mime = event.mimeData()
        if mime.hasUrls():
            for url in mime.urls():
                if not url.isLocalFile():
                    continue
                path = Path(url.toLocalFile())
                if not path.is_file():
                    continue
                try:
                    text = path.read_text(encoding="utf-8", errors="replace")
                except OSError as exc:
                    self.fileDropError.emit(f"Could not read {path.name}: {exc}")
                    event.acceptProposedAction()
                    return
                self.fileDropped.emit(text)
                event.acceptProposedAction()
                return
            event.ignore()
            return
        if mime.hasText():
            self.fileDropped.emit(mime.text())
            event.acceptProposedAction()
            return
        super().dropEvent(event)


class DropRefDesTable(QTableWidget):
    fileDropped = Signal(str)

    def __init__(self, rows: int, columns: int) -> None:
        super().__init__(rows, columns)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event) -> None:  # noqa: N802
        if _first_supported_refdes_change_path(event.mimeData()) is not None:
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dragMoveEvent(self, event) -> None:  # noqa: N802
        if _first_supported_refdes_change_path(event.mimeData()) is not None:
            event.acceptProposedAction()
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event) -> None:  # noqa: N802
        path = _first_supported_refdes_change_path(event.mimeData())
        if path is not None:
            self.fileDropped.emit(str(path))
            event.acceptProposedAction()
            return
        super().dropEvent(event)


@dataclass(frozen=True)
class RefDesComponentChange:
    refdes_name: str
    old_component: str
    new_component: str


@dataclass(frozen=True)
class RefDesComponentChangeBatch:
    changes: list[RefDesComponentChange]


VALID_ACTIVATION_STATUSES = ("Enabled", "Disabled", "Automatic")
IMPORTABLE_ACTIVATION_STATUSES = (*VALID_ACTIVATION_STATUSES, "Unknown")
LEGACY_REFDES_STATUS_HEADER = ("Component", "RefDes Name", "Activation Status")
LEGACY_REFDES_STATUS_EXPORT_HEADER = (*LEGACY_REFDES_STATUS_HEADER, "Net Name")
PARTIAL_REFDES_STATUS_HEADER = ("Component type", "REFDES", "Status")
REFDES_STATUS_HEADER = ("RefDes", "Status")
REFDES_STATUS_HEADERS = (
    LEGACY_REFDES_STATUS_HEADER,
    LEGACY_REFDES_STATUS_EXPORT_HEADER,
    PARTIAL_REFDES_STATUS_HEADER,
    REFDES_STATUS_HEADER,
)
IMPORTABLE_ACTIVATION_STATUS_TOKENS = frozenset(status.casefold() for status in IMPORTABLE_ACTIVATION_STATUSES)
COMPONENT_HEADER_ALIASES = frozenset({"component", "component type", "component name"})
REFDES_HEADER_ALIASES = frozenset({"refdes", "refdes name"})
STATUS_HEADER_ALIASES = frozenset({"status", "activation status"})
NET_HEADER_ALIASES = frozenset({"net", "net name"})


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(f"SPD Model Injector {__version__}")
        self.resize(1400, 760)
        self.setMinimumSize(900, 600)

        self.spd_path: Path | None = None
        self._pending_spd_path: Path | None = None
        self.inventory = SpdInventory(blocks=[], refdes_records=[])
        self.blocks: list[PartialCktBlock] = []
        self.refdes_records: list[RefDesRecord] = []
        self.refdes_by_component: dict[str, list[RefDesRecord]] = {}
        self.refdes_component_changes: dict[str, str] = {}
        self.refdes_activation_status_changes: dict[str, str] = {}
        self.refdes_component_undo_stack: list[RefDesComponentChangeBatch] = []
        self.replacements: dict[str, str] = {}
        self.component_renames: dict[str, str] = {}
        self.component_clones: dict[str, str] = {}
        self._loading_editor = False
        self._busy = False
        self._scan_thread: QThread | None = None
        self._scan_worker: ScanWorker | None = None
        self._export_thread: QThread | None = None
        self._export_worker: ExportWorker | None = None
        self.export_refdes_action: QAction | None = None
        self.import_refdes_status_action: QAction | None = None
        self.undo_component_change_action: QAction | None = None

        self.component_list_label = QLabel("PartialCkt Components (0/0)")
        self.component_filter = QLineEdit()
        self.component_filter.setPlaceholderText("Filter components...")
        self.component_filter.setToolTip("Filter components by name (Ctrl+F)")
        self.component_filter.textChanged.connect(self._apply_component_filter)

        self.component_list = QListWidget()
        self.component_list.currentRowChanged.connect(self._on_current_row_changed)
        self.component_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.component_list.customContextMenuRequested.connect(self._show_component_context_menu)

        self.component_label = QLabel("No component selected")
        self.mapping_label = QLabel("Load an SPD file to inspect PartialCkt blocks.")
        self.mapping_label.setWordWrap(True)

        self.refdes_label = QLabel("Load an SPD file to inspect RefDes records.")
        self.refdes_label.setWordWrap(True)
        self.refdes_table = DropRefDesTable(0, 2)
        self.refdes_table.setHorizontalHeaderLabels(["RefDes Name", "Activation Status"])
        self.refdes_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.refdes_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.refdes_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.refdes_table.setSortingEnabled(True)
        self.refdes_table.setAlternatingRowColors(True)
        self.refdes_table.setShowGrid(True)
        self.refdes_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.refdes_table.customContextMenuRequested.connect(self._show_refdes_context_menu)
        self.refdes_table.fileDropped.connect(self.import_refdes_drop_file)
        refdes_header = self.refdes_table.horizontalHeader()
        refdes_header.setSectionsClickable(True)
        refdes_header.setSortIndicatorShown(True)
        refdes_header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        refdes_header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        refdes_header.resizeSection(0, 180)
        refdes_header.resizeSection(1, 130)
        refdes_header.setSortIndicator(0, Qt.SortOrder.AscendingOrder)

        self.editor = DropModelEditor()
        self.editor.fileDropped.connect(self.import_model_text)
        self.editor.fileDropError.connect(lambda message: self._show_validation(message, error=True))
        self.editor.textChanged.connect(self._on_editor_changed)

        self.validation_label = QLabel("")
        self.validation_label.setWordWrap(True)
        self.status_label = QLabel("Load an SPD file to begin.")
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setRange(0, 100)
        self.progress.setValue(0)

        self.status_log = QPlainTextEdit()
        self.status_log.setReadOnly(True)
        self.status_log.setMaximumBlockCount(200)
        self.status_log.setFixedHeight(120)
        self.status_log.setPlaceholderText("Status details will appear here.")

        self._build_layout()
        self._build_toolbar()
        self._build_menus()
        self._build_shortcuts()
        self.setStatusBar(QStatusBar())
        self.statusBar().addWidget(self.status_label, 1)
        self.statusBar().addPermanentWidget(self.progress)

    def _build_toolbar(self) -> None:
        toolbar = QToolBar("Main")
        toolbar.setMovable(False)
        self.toolBar = toolbar
        self.addToolBar(toolbar)

        self.load_action = QAction("Load SPD", self)
        self.load_action.setShortcut(QKeySequence("Ctrl+O"))
        self.load_action.setToolTip("Load an SPD file and scan for PartialCkt blocks (Ctrl+O)")
        self.load_action.triggered.connect(self.load_spd_dialog)
        toolbar.addAction(self.load_action)

        self.validate_action = QAction("Validate/Convert", self)
        self.validate_action.setShortcut(QKeySequence("Ctrl+Return"))
        self.validate_action.setToolTip("Validate or convert the current editor text (Ctrl+Return)")
        self.validate_action.triggered.connect(self.validate_current_text)
        toolbar.addAction(self.validate_action)

        self.revert_action = QAction("Revert", self)
        self.revert_action.setToolTip("Revert the selected component to the body stored in the SPD")
        self.revert_action.triggered.connect(self.revert_current)
        toolbar.addAction(self.revert_action)

        self.export_action = QAction("Export New SPD", self)
        self.export_action.setShortcut(QKeySequence("Ctrl+E"))
        self.export_action.setToolTip("Export a new SPD with modified bodies replaced (Ctrl+E)")
        self.export_action.triggered.connect(self.export_spd_dialog)
        toolbar.addAction(self.export_action)

        self.export_refdes_action = QAction("Export RefDes Excel", self)
        self.export_refdes_action.setEnabled(False)
        self.export_refdes_action.triggered.connect(self.export_refdes_dialog)
        toolbar.addAction(self.export_refdes_action)

        self.import_refdes_status_action = QAction("Import RefDes Excel", self)
        self.import_refdes_status_action.setEnabled(False)
        self.import_refdes_status_action.triggered.connect(self.import_refdes_dialog)
        toolbar.addAction(self.import_refdes_status_action)

    def _build_menus(self) -> None:
        edit_menu = self.menuBar().addMenu("Edit")
        self.undo_component_change_action = QAction("Undo Component Change", self)
        self.undo_component_change_action.setShortcut(QKeySequence.StandardKey.Undo)
        self.undo_component_change_action.setShortcutContext(Qt.ShortcutContext.ApplicationShortcut)
        self.undo_component_change_action.setEnabled(False)
        self.undo_component_change_action.triggered.connect(self.undo_refdes_component_change)
        edit_menu.addAction(self.undo_component_change_action)
        self.help_menu = self.menuBar().addMenu("Help")
        formats_action = QAction("Input File Formats", self)
        formats_action.triggered.connect(self.show_input_file_formats)
        self.help_menu.addAction(formats_action)

    def _build_layout(self) -> None:
        root = QSplitter(Qt.Orientation.Horizontal)
        root.setObjectName("root_splitter")

        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.addWidget(self.component_list_label)
        left_layout.addWidget(self.component_filter)
        left_layout.addWidget(self.component_list)
        root.addWidget(left)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        work_splitter = QSplitter(Qt.Orientation.Horizontal)
        work_splitter.setObjectName("work_splitter")

        center = QWidget()
        center_layout = QVBoxLayout(center)
        center_layout.addWidget(self.component_label)
        center_layout.addWidget(self.mapping_label)

        button_row = QHBoxLayout()
        self.import_button = QPushButton("Import .mod/.txt")
        self.import_button.setToolTip("Import a SPICE model file into the editor (Ctrl+I)")
        self.import_button.clicked.connect(self.import_model_dialog)
        self.validate_button = QPushButton("Validate/Convert")
        self.validate_button.setToolTip("Validate or convert the current editor text (Ctrl+Return)")
        self.validate_button.clicked.connect(self.validate_current_text)
        self.revert_button = QPushButton("Revert")
        self.revert_button.setToolTip("Revert the selected component to the body stored in the SPD")
        self.revert_button.clicked.connect(self.revert_current)
        button_row.addWidget(self.import_button)
        button_row.addWidget(self.validate_button)
        button_row.addWidget(self.revert_button)
        button_row.addStretch(1)
        center_layout.addLayout(button_row)

        center_layout.addWidget(self.editor, 1)
        center_layout.addWidget(self.validation_label)
        work_splitter.addWidget(center)

        self.refdes_panel = QFrame()
        self.refdes_panel.setObjectName("refdes_panel")
        self.refdes_panel.setFrameShape(QFrame.Shape.StyledPanel)
        self.refdes_panel.setFrameShadow(QFrame.Shadow.Sunken)
        self.refdes_panel.setLineWidth(1)
        refdes_layout = QVBoxLayout(self.refdes_panel)
        refdes_layout.setContentsMargins(8, 8, 8, 8)
        refdes_layout.setSpacing(6)
        refdes_layout.addWidget(QLabel("RefDes List"))
        refdes_layout.addWidget(self.refdes_label)
        refdes_layout.addWidget(self.refdes_table, 1)
        work_splitter.addWidget(self.refdes_panel)
        work_splitter.setSizes([720, 380])

        right_layout.addWidget(work_splitter, 1)
        right_layout.addWidget(QLabel("Status"))
        right_layout.addWidget(self.status_log)
        root.addWidget(right)
        root.setSizes([300, 1100])
        self.setCentralWidget(root)

    def _build_shortcuts(self) -> None:
        self._import_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
        self._import_shortcut.activated.connect(self.import_model_dialog)

        self._filter_shortcut = QShortcut(QKeySequence("Ctrl+F"), self)
        self._filter_shortcut.activated.connect(self._focus_component_filter)

    def _focus_component_filter(self) -> None:
        self.component_filter.setFocus()
        self.component_filter.selectAll()

    def _set_busy(self, busy: bool) -> None:
        self._busy = busy
        for action in (self.load_action, self.validate_action, self.export_action):
            action.setEnabled(not busy)
        if self.import_refdes_status_action is not None:
            self.import_refdes_status_action.setEnabled(not busy and bool(self.refdes_records))
        for button in (self.import_button, self.validate_button):
            button.setEnabled(not busy)

    def _clear_scan_refs(self) -> None:
        if self.sender() is self._scan_thread:
            self._scan_thread = None
            self._scan_worker = None

    def _clear_export_refs(self) -> None:
        if self.sender() is self._export_thread:
            self._export_thread = None
            self._export_worker = None

    def closeEvent(self, event) -> None:  # noqa: N802
        for thread in (self._scan_thread, self._export_thread):
            if thread is not None and thread.isRunning():
                thread.quit()
                thread.wait(5000)
        super().closeEvent(event)

    def load_spd_dialog(self) -> None:
        directory = str(self.spd_path.parent) if self.spd_path else ""
        path, _ = QFileDialog.getOpenFileName(self, "Load SPD", directory, "SPD files (*.spd);;All files (*)")
        if path:
            self.load_spd(path)

    def load_spd(self, path: str | Path) -> None:
        pending_path = Path(path)
        self._pending_spd_path = pending_path
        self.inventory = SpdInventory(blocks=[], refdes_records=[])
        self.blocks = []
        self.refdes_records = []
        self.refdes_by_component = {}
        self.refdes_component_changes.clear()
        self.refdes_activation_status_changes.clear()
        self.refdes_component_undo_stack.clear()
        self.replacements.clear()
        self.component_renames.clear()
        self.component_clones.clear()
        self.component_list.clear()
        if self.component_filter.text():
            self.component_filter.blockSignals(True)
            self.component_filter.clear()
            self.component_filter.blockSignals(False)
        self._update_component_header()
        self._update_undo_component_change_action()
        if self.export_refdes_action is not None:
            self.export_refdes_action.setEnabled(False)
        if self.import_refdes_status_action is not None:
            self.import_refdes_status_action.setEnabled(False)
        self._populate_refdes_table(None)
        self.component_label.setText("No component selected")
        self.mapping_label.setText("Scanning for PartialCkt blocks...")
        self._set_editor_text("")
        self._show_validation("", error=False)
        self.status_log.clear()
        self._append_status(f"Load requested: {pending_path}")

        self.status_label.setText(f"Scanning {pending_path.name}...")
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setVisible(True)
        self._set_busy(True)
        self._scan_thread = QThread(self)
        self._scan_worker = ScanWorker(pending_path)
        self._scan_worker.moveToThread(self._scan_thread)
        self._scan_thread.started.connect(self._scan_worker.run)
        self._scan_worker.progress.connect(self._scan_progress)
        self._scan_worker.finished.connect(self._scan_finished)
        self._scan_worker.failed.connect(self._scan_failed)
        self._scan_worker.finished.connect(self._scan_thread.quit)
        self._scan_worker.failed.connect(self._scan_thread.quit)
        self._scan_thread.finished.connect(self._scan_worker.deleteLater)
        self._scan_thread.finished.connect(self._clear_scan_refs)
        self._scan_thread.start()

    def _scan_progress(self, message: str, current: int, total: int) -> None:
        if total > 0:
            percent = min(100, max(0, int(current * 100 / total)))
            self.progress.setValue(percent)
            detail = f"{message} ({percent}%, {_format_bytes(current)} / {_format_bytes(total)})"
        else:
            self.progress.setValue(0)
            detail = message
        self.status_label.setText(detail)
        self._append_status(detail)

    def _scan_finished(self, inventory: SpdInventory) -> None:
        self.spd_path = self._pending_spd_path
        self._pending_spd_path = None
        self.inventory = inventory
        self.blocks = inventory.blocks
        self.refdes_records = inventory.refdes_records
        self.refdes_component_changes.clear()
        self.refdes_activation_status_changes.clear()
        self.refdes_component_undo_stack.clear()
        self.rebuild_refdes_groups()
        self.replacements.clear()
        self.component_renames.clear()
        self.component_clones.clear()
        self.populate_components()
        self._update_undo_component_change_action()
        if self.export_refdes_action is not None:
            self.export_refdes_action.setEnabled(bool(self.refdes_records))
        if self.import_refdes_status_action is not None:
            self.import_refdes_status_action.setEnabled(bool(self.refdes_records))
        self.progress.setValue(100)
        self.progress.setVisible(False)
        self._set_busy(False)
        message = f"Loaded {len(self.blocks)} PartialCkt blocks and {len(self.refdes_records)} RefDes records."
        self.status_label.setText(message)
        self._append_status(message)
        if self.blocks:
            self.component_list.setCurrentRow(0)
            self._on_current_row_changed(self.component_list.currentRow())
        else:
            self._populate_refdes_table(None)

    def _scan_failed(self, message: str) -> None:
        self._pending_spd_path = None
        self.spd_path = None
        self.inventory = SpdInventory(blocks=[], refdes_records=[])
        self.blocks = []
        self.refdes_records = []
        self.refdes_by_component = {}
        self.refdes_component_changes.clear()
        self.refdes_activation_status_changes.clear()
        self.refdes_component_undo_stack.clear()
        self.replacements.clear()
        self.component_renames.clear()
        self.component_clones.clear()
        self.component_list.clear()
        self._update_component_header()
        self._update_undo_component_change_action()
        if self.import_refdes_status_action is not None:
            self.import_refdes_status_action.setEnabled(False)
        self._populate_refdes_table(None)
        self.component_label.setText("No component selected")
        self.mapping_label.setText("Load an SPD file to inspect PartialCkt blocks.")
        self._set_editor_text("")
        self._show_worker_error(message)

    def populate_components(self) -> None:
        self.component_list.clear()
        for block in self.blocks:
            item = QListWidgetItem()
            item.setData(Qt.ItemDataRole.UserRole, block.component_name)
            self.component_list.addItem(item)
            self._style_item(item, block)
        if self.component_filter.text():
            self.component_filter.blockSignals(True)
            self.component_filter.clear()
            self.component_filter.blockSignals(False)
        self._update_component_header()

    def _apply_component_filter(self, text: str) -> None:
        needle = text.strip().lower()
        for row in range(self.component_list.count()):
            item = self.component_list.item(row)
            name = str(item.data(Qt.ItemDataRole.UserRole) or "")
            item.setHidden(bool(needle) and needle not in name.lower())
        self._update_component_header()

    def _update_component_header(self) -> None:
        total = self.component_list.count()
        visible = sum(1 for row in range(total) if not self.component_list.item(row).isHidden())
        self.component_list_label.setText(f"PartialCkt Components ({visible}/{total})")

    def import_model_dialog(self) -> None:
        directory = str(self.spd_path.parent) if self.spd_path else ""
        path, _ = QFileDialog.getOpenFileName(
            self, "Import SPICE Model", directory, "Model files (*.mod *.txt);;All files (*)"
        )
        if not path:
            return
        try:
            text = Path(path).read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            self._show_validation(f"Could not read {Path(path).name}: {exc}", error=True)
            return
        self.import_model_text(text)

    def import_model_text(self, raw_model: str) -> None:
        block = self.current_block()
        if block is None:
            self._show_validation("Select a PartialCkt before importing a model.", error=True)
            return
        try:
            prepared = prepare_model_for_partialckt(raw_model, block.ext_nodes)
        except ModelValidationError as exc:
            self._show_validation(str(exc), error=True)
            return
        self._set_editor_text(prepared)
        self.replacements[block.component_name] = prepared
        self._show_mapping(block, raw_model)
        self._show_validation(f"OK: {block.port_count} model ports matched {block.port_count} ExtNode ports.", error=False)
        self._refresh_current_item()

    def validate_current_text(self) -> None:
        text = self.editor.toPlainText()
        if ".SUBCKT" in text.upper():
            self.import_model_text(text)
            return
        block = self.current_block()
        if block is None:
            self._show_validation("No component selected.", error=True)
            return
        self.replacements[block.component_name] = _ensure_lf_ending(text)
        self._show_validation("OK: manual PartialCkt body will be used as edited.", error=False)
        self._refresh_current_item()

    def revert_current(self) -> None:
        block = self.current_block()
        if block is None:
            return
        self.replacements.pop(block.component_name, None)
        self._load_block_body(block)
        self._show_validation("Reverted to the body currently stored in the SPD.", error=False)
        self._refresh_current_item()

    def export_spd_dialog(self) -> None:
        if self.spd_path is None:
            QMessageBox.warning(self, "Export", "Load an SPD file first.")
            return
        default_path = str(self.spd_path.parent / f"{self.spd_path.stem}_injected.spd")
        path, _ = QFileDialog.getSaveFileName(self, "Export New SPD", default_path, "SPD files (*.spd);;All files (*)")
        if not path:
            return

        output_path = Path(path)
        try:
            same_as_source = output_path.resolve() == self.spd_path.resolve()
        except OSError:
            same_as_source = False
        if same_as_source or output_path == self.spd_path:
            QMessageBox.warning(
                self,
                "Export",
                "Output path must differ from the source SPD file. Choose a different filename.",
            )
            return

        if not self.replacements and not self.refdes_component_changes and not self.refdes_activation_status_changes and not self.component_renames and not self.component_clones:
            confirm = QMessageBox.question(
                self,
                "Export",
                "No components are modified; the output will be an LF-normalized copy of the source. Continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm != QMessageBox.StandardButton.Yes:
                return

        self.export_spd(output_path)

    def export_refdes_dialog(self) -> None:
        if not self.refdes_records:
            QMessageBox.information(self, "Export RefDes Excel", "No RefDes records are loaded.")
            return
        default_path = "refdes.xlsx"
        if self.spd_path is not None:
            default_path = str(self.spd_path.with_name(f"{self.spd_path.stem}_refdes.xlsx"))
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export RefDes Excel",
            default_path,
            "Excel files (*.xlsx);;All files (*)",
        )
        if path:
            self.export_refdes_excel(path)

    def export_refdes_excel(self, output_path: str | Path) -> None:
        export_refdes_xlsx(output_path, self.effective_refdes_records())
        message = f"Exported RefDes Excel: {output_path}"
        self.status_label.setText(message)
        self._append_status(message)

    def import_refdes_dialog(self) -> None:
        if not self.refdes_records:
            QMessageBox.information(self, "Import RefDes Excel", "No RefDes records are loaded.")
            return
        directory = str(self.spd_path.parent) if self.spd_path else ""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import RefDes Excel",
            directory,
            "RefDes files (*.xlsx *.csv);;All files (*)",
        )
        if path:
            self.import_refdes_file(path)

    # Compatibility for callers using the pre-0.1.10 method name.
    def import_refdes_status_dialog(self) -> None:
        self.import_refdes_dialog()

    def show_input_file_formats(self) -> None:
        QMessageBox.information(
            self,
            "Input File Formats",
            "Accepted RefDes files (XLSX or CSV):\n"
            "• Component changes: RefDes | Component (optional header).\n"
            "• Status updates: RefDes | Status; Component | RefDes | Status; "
            "or Component | RefDes | Status | Net (Net is ignored).\n"
            "Headers are optional, and header names are case-insensitive with surrounding whitespace ignored; aliases include "
            "Component/Component type/Component Name, RefDes/RefDes Name, "
            "Status/Activation Status, and Net/Net Name.\n"
            "Status cells must be exactly Enabled, Disabled, Automatic, or Unknown (Unknown is valid only when already present). "
            "Menu import and RefDes-table drag-and-drop use identical auto-detection.",
        )

    def export_spd(self, output_path: str | Path) -> None:
        if self.spd_path is None:
            return
        output_path = Path(output_path)
        self.status_label.setText(f"Writing {output_path.name}...")
        self._append_status(f"Writing SPD output: {output_path}")
        self.progress.setRange(0, 0)
        self.progress.setVisible(True)
        self._set_busy(True)
        self._export_thread = QThread(self)
        self._export_worker = ExportWorker(
            self.spd_path,
            output_path,
            self.blocks,
            dict(self.replacements),
            dict(self.refdes_component_changes),
            dict(self.refdes_activation_status_changes),
            list(self.refdes_records),
            dict(self.component_renames),
            dict(self.component_clones),
        )
        self._export_worker.moveToThread(self._export_thread)
        self._export_thread.started.connect(self._export_worker.run)
        self._export_worker.finished.connect(self._export_finished)
        self._export_worker.failed.connect(self._show_worker_error)
        self._export_worker.finished.connect(self._export_thread.quit)
        self._export_worker.failed.connect(self._export_thread.quit)
        self._export_thread.finished.connect(self._export_worker.deleteLater)
        self._export_thread.finished.connect(self._clear_export_refs)
        self._export_thread.start()

    def _export_finished(self, output_path: str) -> None:
        self.progress.setVisible(False)
        self._set_busy(False)
        message = f"Exported to {output_path}"
        self.status_label.setText(message)
        self._append_status(message)

    def _show_worker_error(self, message: str) -> None:
        self.progress.setVisible(False)
        self._set_busy(False)
        self.status_label.setText("Operation failed.")
        self._append_status(f"Operation failed: {message}")
        box = QMessageBox(
            QMessageBox.Icon.Critical,
            "SPD Model Injector",
            message,
            QMessageBox.StandardButton.Ok,
            self,
        )
        box.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        box.open()

    def _on_current_row_changed(self, row: int) -> None:
        if row < 0 or row >= len(self.blocks):
            return
        block = self.blocks[row]
        self.component_label.setText(block.component_name)
        self._show_mapping(block)
        self._populate_refdes_table(block.component_name)
        self._append_status(f"Selected {block.component_name}: {block.port_count} ports, lines {block.start_line}-{block.end_line}")
        if block.component_name in self.replacements:
            self._set_editor_text(self.replacements[block.component_name])
        else:
            self._load_block_body(block)

    def _on_editor_changed(self) -> None:
        if self._loading_editor:
            return
        block = self.current_block()
        if block is None:
            return
        self.replacements[block.component_name] = _ensure_lf_ending(self.editor.toPlainText())
        self._show_validation("Edited manually. Use Validate/Convert before export if this is a full .SUBCKT model.", error=False)
        self._refresh_current_item()

    def _load_block_body(self, block: PartialCktBlock) -> None:
        if self.spd_path is None:
            self._set_editor_text("")
            return
        try:
            body = read_block_body(self.spd_path, block)
        except OSError as exc:
            self._set_editor_text("")
            message = f"Could not read {self.spd_path.name}: {exc}"
            self._show_validation(message, error=True)
            self.status_label.setText(message)
            self._append_status(message)
            return
        self._set_editor_text(body)
        self.status_label.setText(f"Loaded body for {block.component_name}.")

    def _set_editor_text(self, text: str) -> None:
        self._loading_editor = True
        self.editor.setPlainText(text)
        self._loading_editor = False

    def current_block(self) -> PartialCktBlock | None:
        row = self.component_list.currentRow()
        if row < 0 or row >= len(self.blocks):
            return None
        return self.blocks[row]

    def _show_mapping(self, block: PartialCktBlock, raw_model: str | None = None) -> None:
        model_info = ""
        if raw_model and ".SUBCKT" in raw_model.upper():
            first_line = next((line for line in raw_model.splitlines() if line.strip().upper().startswith(".SUBCKT")), "")
            model_info = f" | Model: {first_line.strip()}"
        ext_preview = " ".join(block.ext_nodes[:20])
        if len(block.ext_nodes) > 20:
            ext_preview += f" ... (+{len(block.ext_nodes) - 20} more)"
        self.mapping_label.setText(f"ExtNode ({block.port_count}): {ext_preview}{model_info}")

    def _is_dark_palette(self) -> bool:
        return self.palette().color(QPalette.ColorRole.Window).lightness() < 128

    def _modified_color(self) -> QColor:
        return QColor("#fbbf24") if self._is_dark_palette() else QColor("#b45309")

    def _show_validation(self, message: str, *, error: bool) -> None:
        if not message:
            self.validation_label.setText("")
            return
        dark = self._is_dark_palette()
        if error:
            color = "#f87171" if dark else "#b91c1c"
            style = f"color:{color};font-weight:bold;"
        else:
            color = "#34d399" if dark else "#065f46"
            style = f"color:{color};"
        escaped = html.escape(message)
        self.validation_label.setText(f"<span style='{style}'>{escaped}</span>")

    def _populate_refdes_table(self, component_name: str | None) -> None:
        self.refdes_table.setRowCount(0)
        if component_name is None:
            self.refdes_label.setText("Select a component to inspect RefDes records.")
            return

        records = self.refdes_by_component.get(component_name, [])
        self.refdes_label.setText(f"{component_name}: {len(records)} RefDes records")
        sorting_enabled = self.refdes_table.isSortingEnabled()
        self.refdes_table.setSortingEnabled(False)
        self.refdes_table.setRowCount(len(records))
        for row, record in enumerate(records):
            refdes_item = QTableWidgetItem(record.refdes_name)
            refdes_item.setData(Qt.ItemDataRole.UserRole, record.refdes_name)
            self.refdes_table.setItem(row, 0, refdes_item)
            self.refdes_table.setItem(row, 1, QTableWidgetItem(record.activation_status))
        self.refdes_table.setSortingEnabled(sorting_enabled)

    def _show_refdes_context_menu(self, position) -> None:
        row = self.refdes_table.indexAt(position).row()
        if row < 0:
            return
        if not self.refdes_table.selectionModel().isRowSelected(row, self.refdes_table.rootIndex()):
            self.refdes_table.clearSelection()
            self.refdes_table.selectRow(row)
        refdes_names = self.selected_refdes_names()
        if not refdes_names:
            return
        menu = QMenu(self.refdes_table)
        change_action = menu.addAction(f"Change Component... ({len(refdes_names)} RefDes)")
        menu.addSeparator()
        status_actions = {
            menu.addAction(f"Set Status: {status}"): status
            for status in VALID_ACTIVATION_STATUSES
        }
        selected_action = menu.exec(self.refdes_table.viewport().mapToGlobal(position))
        if selected_action is change_action:
            self.change_selected_refdes_components(refdes_names)
        elif selected_action in status_actions:
            self.apply_refdes_activation_status_changes(refdes_names, status_actions[selected_action])

    def _show_component_context_menu(self, position) -> None:
        if self._busy:
            return
        row = self.component_list.indexAt(position).row()
        if row < 0 or row >= len(self.blocks):
            return
        self.component_list.setCurrentRow(row)
        menu = QMenu(self.component_list)
        clone_action = menu.addAction("Clone Component...")
        rename_action = menu.addAction("Rename Component...")
        selected = menu.exec(self.component_list.viewport().mapToGlobal(position))
        if selected is clone_action:
            self.clone_current_component()
        elif selected is rename_action:
            self.rename_current_component()

    def _prompt_component_name(self, title: str, current: str = "") -> str | None:
        existing = {block.component_name.casefold() for block in self.blocks}
        while True:
            name, accepted = QInputDialog.getText(self, title, "Component Name", text=current)
            if not accepted:
                return None
            name = name.strip()
            if not name or any(char.isspace() for char in name):
                QMessageBox.warning(self, title, "Component name must be a non-empty single token.")
                continue
            if name.casefold() in existing and name.casefold() != current.casefold():
                QMessageBox.warning(self, title, f"Component already exists: {name}")
                continue
            return name

    def clone_current_component(self) -> None:
        if self._busy:
            return
        source = self.current_block()
        if source is None:
            return
        name = self._prompt_component_name("Clone Component")
        if not name:
            return
        source_name = source.component_name
        source_origin = source.source_component_name or source.clone_source_name or source_name
        clone = replace(source, component_name=name, clone_source_name=source_origin)
        self.blocks.append(clone)
        self.component_clones[name] = source_origin
        if source_name in self.replacements:
            self.replacements[name] = self.replacements[source_name]
        elif self.spd_path is not None:
            self.replacements[name] = read_block_body(self.spd_path, source)
        item = QListWidgetItem()
        item.setData(Qt.ItemDataRole.UserRole, name)
        self.component_list.addItem(item)
        self._style_item(item, clone)
        self._update_component_header()
        self.component_list.setCurrentRow(self.component_list.count() - 1)
        self._append_status(f"Cloned {source_name} as {name}.")

    def rename_current_component(self) -> None:
        if self._busy:
            return
        block = self.current_block()
        if block is None:
            return
        row = self.component_list.currentRow()
        old_name = block.component_name
        name = self._prompt_component_name("Rename Component", old_name)
        if not name or name == old_name:
            return
        source_name = block.source_component_name
        if old_name in self.component_clones:
            self.component_clones[name] = self.component_clones.pop(old_name)
        else:
            origin = source_name or old_name
            if name == origin:
                self.component_renames.pop(origin, None)
            else:
                self.component_renames[origin] = name
        header_origin = source_name if block.clone_source_name is not None else (source_name or old_name)
        self.blocks[row] = replace(block, component_name=name, source_component_name=header_origin)
        if old_name in self.replacements:
            self.replacements[name] = self.replacements.pop(old_name)
        # Keep assignments coherent and invalidate undo batches containing the old identity.
        for record in self.refdes_records:
            if self.effective_component_for_refdes(record.refdes_name) == old_name:
                self._set_refdes_effective_component(record.refdes_name, name)
        self.refdes_component_undo_stack.clear()
        self.populate_components()
        self.component_list.setCurrentRow(row)
        self.rebuild_refdes_groups()
        self._populate_refdes_table(name)
        self._update_undo_component_change_action()
        self._append_status(f"Renamed {old_name} to {name}.")

    def change_selected_refdes_components(self, refdes_names: list[str] | None = None) -> None:
        names = refdes_names or self.selected_refdes_names()
        if not names:
            return
        component_name = self._choose_component_name()
        if component_name:
            self.apply_refdes_component_changes(names, component_name)

    def _choose_component_name(self) -> str | None:
        components = [block.component_name for block in self.blocks]
        if not components:
            QMessageBox.warning(self, "Change Component", "No components are loaded.")
            return None

        dialog = QDialog(self)
        dialog.setWindowTitle("Change Component")
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Component Name"))
        combo = QComboBox(dialog)
        combo.setEditable(True)
        combo.addItems(components)
        completer = QCompleter(components, combo)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        combo.setCompleter(completer)
        layout.addWidget(combo)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, dialog)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return None
        component_name = combo.currentText().strip()
        if component_name not in components:
            QMessageBox.warning(self, "Change Component", f"Unknown Component: {component_name}")
            return None
        return component_name

    def selected_refdes_names(self) -> list[str]:
        names: list[str] = []
        seen: set[str] = set()
        for index in self.refdes_table.selectionModel().selectedRows(0):
            item = self.refdes_table.item(index.row(), 0)
            if item is None:
                continue
            refdes_name = item.data(Qt.ItemDataRole.UserRole) or item.text()
            if refdes_name not in seen:
                seen.add(refdes_name)
                names.append(refdes_name)
        return names

    def apply_refdes_component_changes(self, refdes_names: list[str], component_name: str) -> None:
        errors = self.validate_refdes_component_changes({name: component_name for name in refdes_names})
        if errors:
            QMessageBox.warning(self, "Change Component", "\n".join(errors))
            return

        changes: list[RefDesComponentChange] = []
        for refdes_name in refdes_names:
            old_component = self.effective_component_for_refdes(refdes_name)
            if old_component is None or old_component == component_name:
                continue
            changes.append(RefDesComponentChange(refdes_name, old_component, component_name))
        if not changes:
            return

        self.refdes_component_undo_stack.append(RefDesComponentChangeBatch(changes))
        self.refdes_component_undo_stack = self.refdes_component_undo_stack[-10:]
        for change in changes:
            self._set_refdes_effective_component(change.refdes_name, change.new_component)
        self._after_refdes_component_change(f"Changed {len(changes)} RefDes component assignment(s).")

    def undo_refdes_component_change(self) -> None:
        if not self.refdes_component_undo_stack:
            return
        batch = self.refdes_component_undo_stack.pop()
        for change in batch.changes:
            self._set_refdes_effective_component(change.refdes_name, change.old_component)
        self._after_refdes_component_change(f"Undid {len(batch.changes)} RefDes component assignment(s).")

    def _after_refdes_component_change(self, message: str) -> None:
        self.rebuild_refdes_groups()
        current = self.current_block()
        self._populate_refdes_table(current.component_name if current is not None else None)
        self._update_undo_component_change_action()
        self.status_label.setText(message)
        self._append_status(message)

    def _after_refdes_status_change(self, message: str) -> None:
        self.rebuild_refdes_groups()
        current = self.current_block()
        self._populate_refdes_table(current.component_name if current is not None else None)
        self.status_label.setText(message)
        self._append_status(message)

    def _set_refdes_effective_component(self, refdes_name: str, component_name: str) -> None:
        original = self.original_component_for_refdes(refdes_name)
        if original is None or original == component_name:
            self.refdes_component_changes.pop(refdes_name, None)
            return
        self.refdes_component_changes[refdes_name] = component_name

    def _set_refdes_effective_activation_status(self, refdes_name: str, activation_status: str) -> None:
        original = self.original_activation_status_for_refdes(refdes_name)
        if original is None or original == activation_status:
            self.refdes_activation_status_changes.pop(refdes_name, None)
            return
        self.refdes_activation_status_changes[refdes_name] = activation_status

    def _update_undo_component_change_action(self) -> None:
        if self.undo_component_change_action is not None:
            self.undo_component_change_action.setEnabled(bool(self.refdes_component_undo_stack))

    def rebuild_refdes_groups(self) -> None:
        grouped: dict[str, list[RefDesRecord]] = {}
        for record in self.effective_refdes_records():
            grouped.setdefault(record.component_name, []).append(record)
        self.refdes_by_component = grouped

    def effective_refdes_records(self) -> list[RefDesRecord]:
        return [
            replace(
                record,
                component_name=self.refdes_component_changes.get(record.refdes_name, record.component_name),
                activation_status=self.refdes_activation_status_changes.get(
                    record.refdes_name,
                    record.activation_status,
                ),
            )
            for record in self.refdes_records
        ]

    def original_component_for_refdes(self, refdes_name: str) -> str | None:
        for record in self.refdes_records:
            if record.refdes_name == refdes_name:
                return record.component_name
        return None

    def effective_component_for_refdes(self, refdes_name: str) -> str | None:
        original = self.original_component_for_refdes(refdes_name)
        if original is None:
            return None
        return self.refdes_component_changes.get(refdes_name, original)

    def original_activation_status_for_refdes(self, refdes_name: str) -> str | None:
        for record in self.refdes_records:
            if record.refdes_name == refdes_name:
                return record.activation_status
        return None

    def effective_activation_status_for_refdes(self, refdes_name: str) -> str | None:
        original = self.original_activation_status_for_refdes(refdes_name)
        if original is None:
            return None
        return self.refdes_activation_status_changes.get(refdes_name, original)

    def apply_refdes_activation_status_changes(self, refdes_names: list[str], activation_status: str) -> None:
        errors = self.validate_refdes_activation_status_changes({name: activation_status for name in refdes_names})
        if errors:
            QMessageBox.warning(self, "Change Activation Status", "\n".join(errors))
            return
        changed = 0
        for refdes_name in refdes_names:
            if self.effective_activation_status_for_refdes(refdes_name) == activation_status:
                continue
            self._set_refdes_effective_activation_status(refdes_name, activation_status)
            changed += 1
        if changed:
            self._after_refdes_status_change(f"Changed {changed} RefDes activation status value(s).")

    def validate_refdes_activation_status_changes(self, changes: dict[str, str]) -> list[str]:
        known_refdes = {record.refdes_name for record in self.refdes_records}
        errors: list[str] = []
        for refdes_name in sorted({name for name in changes if name not in known_refdes}):
            errors.append(f"Unknown RefDes: {refdes_name}")
        for activation_status in sorted({status for status in changes.values() if status not in VALID_ACTIVATION_STATUSES}):
            errors.append(f"Unknown Activation Status: {activation_status}")
        return errors

    def load_refdes_status_file(self, path: str | Path) -> dict[str, str]:
        statuses, _, _, errors = self._parse_refdes_status_file(path)
        if errors:
            raise ValueError("\n".join(errors))
        return statuses

    def validate_refdes_status_file(self, path: str | Path) -> list[str]:
        statuses, components, requires_full_inventory, errors = self._parse_refdes_status_file(path)
        if errors:
            return errors
        return self._validate_imported_refdes_statuses(statuses, components, requires_full_inventory)

    def _validate_imported_refdes_statuses(
        self,
        statuses: dict[str, str],
        components: dict[str, str],
        requires_full_inventory: bool,
    ) -> list[str]:
        current_refdes = {record.refdes_name for record in self.refdes_records}
        imported_refdes = set(statuses)
        validation_errors: list[str] = []
        if requires_full_inventory and len(statuses) != len(self.refdes_records):
            validation_errors.append(f"RefDes count mismatch: expected {len(self.refdes_records)}, got {len(statuses)}")
        if requires_full_inventory:
            for refdes_name in sorted(current_refdes - imported_refdes):
                validation_errors.append(f"Missing RefDes: {refdes_name}")
        for refdes_name in sorted(imported_refdes - current_refdes):
            validation_errors.append(f"Unexpected RefDes: {refdes_name}")
        for refdes_name in sorted(imported_refdes & current_refdes):
            expected_components = {
                component
                for component in (
                    self.original_component_for_refdes(refdes_name),
                    self.effective_component_for_refdes(refdes_name),
                )
                if component is not None
            }
            imported_component = components[refdes_name]
            if imported_component not in expected_components:
                expected = " or ".join(sorted(expected_components))
                validation_errors.append(
                    f"Component mismatch for {refdes_name}: expected {expected}, got {imported_component}"
                )
        for refdes_name, activation_status in statuses.items():
            if refdes_name not in current_refdes:
                continue
            if activation_status in VALID_ACTIVATION_STATUSES:
                continue
            if activation_status == "Unknown" and self.original_activation_status_for_refdes(refdes_name) == "Unknown":
                continue
            validation_errors.append(f"Unknown Activation Status: {activation_status}")
        return validation_errors

    def import_refdes_status_file(self, path: str | Path) -> None:
        statuses, components, requires_full_inventory, errors = self._parse_refdes_status_file(path)
        self._apply_refdes_status_import(statuses, components, requires_full_inventory, errors)

    def import_refdes_file(self, path: str | Path) -> None:
        file_path = Path(path)
        try:
            rows = _read_refdes_rows(file_path)
        except Exception as exc:
            QMessageBox.warning(self, "Import RefDes Excel", f"Could not read RefDes file: {exc}")
            return
        if not rows:
            QMessageBox.warning(self, "Import RefDes Excel", "RefDes file is empty.")
            return
        kind = _refdes_header_kind(rows[0][1])
        if kind == "component":
            self._import_refdes_component_rows(rows)
            return
        if kind in {"status2", "status3", "status4"}:
            self._apply_refdes_status_import(*self._parse_refdes_status_rows(rows))
            return
        if _is_ambiguous_refdes_status_rows(rows, self.refdes_records, self.blocks):
            QMessageBox.warning(
                self,
                "Import RefDes Excel",
                "Ambiguous 2-column RefDes file: values resemble both activation statuses and component names. "
                "Add either a RefDes | Component or RefDes | Status header.",
            )
            return
        logical_count = _headerless_status_column_count(rows)
        if logical_count in (3, 4) or (logical_count == 2 and _looks_like_refdes_status_rows(rows)):
            self._apply_refdes_status_import(*self._parse_refdes_status_rows(rows))
            return
        self._import_refdes_component_rows(rows)

    def _import_refdes_component_rows(self, rows: list[tuple[int, list[str]]]) -> None:
        changes, errors = self._parse_refdes_component_rows(rows)
        if errors:
            QMessageBox.warning(self, "Import RefDes Component Changes", "\n".join(errors))
            return
        self.apply_refdes_component_change_map(changes)

    def _apply_refdes_status_import(
        self,
        statuses: dict[str, str],
        components: dict[str, str],
        requires_full_inventory: bool,
        errors: list[str],
    ) -> None:
        if not errors:
            errors = self._validate_imported_refdes_statuses(statuses, components, requires_full_inventory)
        if errors:
            QMessageBox.warning(self, "Import RefDes Excel", "\n".join(errors))
            return
        changed = 0
        for refdes_name, activation_status in statuses.items():
            if self.effective_activation_status_for_refdes(refdes_name) == activation_status:
                continue
            self._set_refdes_effective_activation_status(refdes_name, activation_status)
            changed += 1
        self._after_refdes_status_change(f"Imported {changed} RefDes activation status value(s).")

    def import_refdes_drop_file(self, path: str | Path) -> None:
        self.import_refdes_file(path)

    def load_refdes_component_change_file(self, path: str | Path) -> dict[str, str]:
        changes, errors = self._parse_refdes_component_change_file(path)
        if errors:
            raise ValueError("\n".join(errors))
        return changes

    def validate_refdes_component_change_file(self, path: str | Path) -> list[str]:
        changes, errors = self._parse_refdes_component_change_file(path)
        if errors:
            return errors
        return self.validate_refdes_component_changes(changes)

    def import_refdes_component_changes_file(self, path: str | Path) -> None:
        errors = self.validate_refdes_component_change_file(path)
        if errors:
            QMessageBox.warning(self, "Import RefDes Component Changes", "\n".join(errors))
            return
        changes = self.load_refdes_component_change_file(path)
        self.apply_refdes_component_change_map(changes)

    def apply_refdes_component_change_map(self, changes: dict[str, str]) -> None:
        errors = self.validate_refdes_component_changes(changes)
        if errors:
            QMessageBox.warning(self, "Import RefDes Component Changes", "\n".join(errors))
            return
        applicable = {
            refdes_name: component_name
            for refdes_name, component_name in changes.items()
            if self.effective_component_for_refdes(refdes_name) != component_name
        }
        batch_changes = [
            RefDesComponentChange(refdes_name, self.effective_component_for_refdes(refdes_name) or "", component_name)
            for refdes_name, component_name in applicable.items()
        ]
        batch_changes = [change for change in batch_changes if change.old_component and change.old_component != change.new_component]
        if not batch_changes:
            return
        self.refdes_component_undo_stack.append(RefDesComponentChangeBatch(batch_changes))
        self.refdes_component_undo_stack = self.refdes_component_undo_stack[-10:]
        for change in batch_changes:
            self._set_refdes_effective_component(change.refdes_name, change.new_component)
        self._after_refdes_component_change(f"Imported {len(batch_changes)} RefDes component assignment(s).")

    def validate_refdes_component_changes(self, changes: dict[str, str]) -> list[str]:
        known_refdes = {record.refdes_name for record in self.refdes_records}
        known_components = {block.component_name for block in self.blocks}
        errors: list[str] = []
        for refdes_name in sorted({name for name in changes if name not in known_refdes}):
            errors.append(f"Unknown RefDes: {refdes_name}")
        for component_name in sorted({name for name in changes.values() if name not in known_components}):
            errors.append(f"Unknown Component: {component_name}")
        return errors

    def _parse_refdes_component_change_file(self, path: str | Path) -> tuple[dict[str, str], list[str]]:
        file_path = Path(path)
        try:
            rows = _read_refdes_rows(file_path)
        except ValueError as exc:
            return {}, [str(exc)]
        except Exception as exc:
            return {}, [f"Could not read RefDes component file: {exc}"]
        return self._parse_refdes_component_rows(rows)

    def _parse_refdes_component_rows(
        self,
        rows: list[tuple[int, list[str]]],
    ) -> tuple[dict[str, str], list[str]]:
        if not rows:
            return {}, ["RefDes component file is empty."]
        if _refdes_header_kind(rows[0][1]) == "component":
            rows = rows[1:]
        if not rows:
            return {}, ["RefDes component file has no data rows."]

        changes: dict[str, str] = {}
        errors: list[str] = []
        for row_number, values in rows:
            refdes_name = values[0].strip() if len(values) > 0 else ""
            component_name = values[1].strip() if len(values) > 1 else ""
            extra_values = [value.strip() for value in values[2:] if value.strip()]
            if not refdes_name or not component_name:
                errors.append(f"Row {row_number}: RefDes and Component Name are required.")
                continue
            if extra_values:
                errors.append(f"Row {row_number}: expected exactly 2 columns.")
                continue
            if refdes_name in changes:
                errors.append(f"Duplicate RefDes: {refdes_name}")
                continue
            changes[refdes_name] = component_name
        return changes, errors

    def _parse_refdes_status_file(
        self,
        path: str | Path,
    ) -> tuple[dict[str, str], dict[str, str], bool, list[str]]:
        file_path = Path(path)
        try:
            rows = _read_refdes_rows(file_path)
        except Exception as exc:
            return {}, {}, False, [f"Could not read RefDes status file: {exc}"]
        return self._parse_refdes_status_rows(rows)

    def _parse_refdes_status_rows(
        self,
        rows: list[tuple[int, list[str]]],
    ) -> tuple[dict[str, str], dict[str, str], bool, list[str]]:
        if not rows:
            return {}, {}, False, ["RefDes status file is empty."]
        header_row, header_values = rows[0]
        recognized_header = _refdes_header_kind(header_values)
        if recognized_header is None:
            # A row whose leading cells look like a header is almost certainly a malformed
            # header, not a headerless data row. Keep the existing actionable error.
            normalized_first_cells = tuple(_normalize_header_value(value) for value in header_values[:4])
            header_like_count = max(
                sum(
                    index < len(expected_header) and value == _normalize_header_value(expected_header[index])
                    for index, value in enumerate(normalized_first_cells)
                )
                for expected_header in REFDES_STATUS_HEADERS
            )
            if header_like_count >= 2:
                expected_headers = " or ".join(", ".join(header) for header in REFDES_STATUS_HEADERS)
                return {}, {}, False, [f"Row {header_row}: expected header: {expected_headers}"]
            column_count = _headerless_status_column_count(rows)
            if column_count not in (2, 3, 4):
                expected_headers = " or ".join(", ".join(header) for header in REFDES_STATUS_HEADERS)
                return {}, {}, False, [f"Row {header_row}: expected header: {expected_headers}"]
            header = None
            requires_full_inventory = False
            has_net_name_column = column_count == 4
            headerless_columns = column_count
        else:
            header = recognized_header
            requires_full_inventory = False
            has_net_name_column = header == "status4"
            headerless_columns = 0
            if len(rows) == 1:
                return {}, {}, requires_full_inventory, ["RefDes status file has no data rows."]

        statuses: dict[str, str] = {}
        components: dict[str, str] = {}
        errors: list[str] = []
        data_rows = rows[1:] if header is not None else rows
        is_two_column = headerless_columns == 2 or header == "status2"
        for row_number, values in data_rows:
            if is_two_column:
                component_name = ""
                refdes_name = values[0].strip() if len(values) > 0 else ""
                activation_status = values[1].strip() if len(values) > 1 else ""
                extra_values = [value.strip() for value in values[2:] if value.strip()]
                inferred_component = self.effective_component_for_refdes(refdes_name)
                if inferred_component is not None:
                    component_name = inferred_component
            else:
                component_name = values[0].strip() if len(values) > 0 else ""
                refdes_name = values[1].strip() if len(values) > 1 else ""
                activation_status = values[2].strip() if len(values) > 2 else ""
                extra_values = [value.strip() for value in (values[4:] if has_net_name_column else values[3:]) if value.strip()]
            if (is_two_column and (not refdes_name or not activation_status)) or (
                not is_two_column and (not component_name or not refdes_name or not activation_status)
            ):
                if is_two_column:
                    errors.append(f"Row {row_number}: RefDes and Activation Status are required.")
                else:
                    errors.append(f"Row {row_number}: Component, RefDes Name, and Activation Status are required.")
                continue
            if extra_values:
                expected_columns = 2 if is_two_column else (4 if has_net_name_column else 3)
                errors.append(f"Row {row_number}: expected exactly {expected_columns} columns.")
                continue
            if refdes_name in statuses:
                errors.append(f"Duplicate RefDes: {refdes_name}")
                continue
            if activation_status not in IMPORTABLE_ACTIVATION_STATUSES:
                errors.append(f"Unknown Activation Status: {activation_status}")
                continue
            statuses[refdes_name] = activation_status
            components[refdes_name] = component_name
        return statuses, components, requires_full_inventory, errors

    def _refresh_current_item(self) -> None:
        block = self.current_block()
        row = self.component_list.currentRow()
        if block is None or row < 0:
            return
        self._style_item(self.component_list.item(row), block)

    def _style_item(self, item: QListWidgetItem, block: PartialCktBlock) -> None:
        item.setText(self._item_text(block))
        modified = block.component_name in self.replacements
        font = item.font()
        font.setBold(modified)
        item.setFont(font)
        if modified:
            item.setForeground(self._modified_color())
        else:
            item.setData(Qt.ItemDataRole.ForegroundRole, None)

    def _item_text(self, block: PartialCktBlock) -> str:
        modified = block.component_name in self.replacements
        renamed = bool(block.source_component_name and block.source_component_name != block.component_name)
        cloned = block.clone_source_name is not None
        marker = "* " if modified or renamed or cloned else ""
        state = "cloned" if cloned else ("renamed" if renamed else ("modified" if modified else "existing"))
        return f"{marker}{block.component_name}\nports: {block.port_count}  {state}"

    def _append_status(self, message: str) -> None:
        self.status_log.appendPlainText(message)


def _ensure_lf_ending(text: str) -> str:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    return normalized if not normalized or normalized.endswith("\n") else normalized + "\n"


def _first_supported_refdes_change_path(mime_data) -> Path | None:
    if not mime_data.hasUrls():
        return None
    for url in mime_data.urls():
        if not url.isLocalFile():
            continue
        path = Path(url.toLocalFile())
        if path.is_file() and path.suffix.lower() in {".csv", ".xlsx"}:
            return path
    return None


def _read_refdes_component_csv(path: Path) -> list[tuple[int, list[str]]]:
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp949")
    rows: list[tuple[int, list[str]]] = []
    for row_number, row in enumerate(csv.reader(io.StringIO(text)), start=1):
        if not any(value.strip() for value in row):
            continue
        rows.append((row_number, row))
    return rows


def _read_refdes_component_xlsx(path: Path) -> list[tuple[int, list[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows: list[tuple[int, list[str]]] = []
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = ["" if value is None else str(value) for value in row]
        if not any(value.strip() for value in values):
            continue
        rows.append((row_number, values))
    workbook.close()
    return rows


def _read_refdes_rows(path: Path) -> list[tuple[int, list[str]]]:
    if path.suffix.lower() == ".csv":
        return _read_refdes_component_csv(path)
    if path.suffix.lower() == ".xlsx":
        return _read_refdes_component_xlsx(path)
    raise ValueError(f"Unsupported file type: {path.suffix}")


def _normalize_header_value(value: str) -> str:
    return value.strip().casefold()


def _logical_column_count(values: list[str]) -> int:
    trimmed = list(values)
    while trimmed and not trimmed[-1].strip():
        trimmed.pop()
    return len(trimmed)


def _headerless_status_column_count(rows: list[tuple[int, list[str]]]) -> int:
    counts = [_logical_column_count(values) for _, values in rows]
    if any(count >= 4 for count in counts):
        return 4
    return counts[0]


def _refdes_header_kind(values: list[str]) -> str | None:
    normalized = [_normalize_header_value(value) for value in values]
    while normalized and not normalized[-1]:
        normalized.pop()
    if len(normalized) == 2:
        if normalized[0] in REFDES_HEADER_ALIASES and normalized[1] in COMPONENT_HEADER_ALIASES:
            return "component"
        if normalized[0] in REFDES_HEADER_ALIASES and normalized[1] in STATUS_HEADER_ALIASES:
            return "status2"
    if len(normalized) == 3 and (
        normalized[0] in COMPONENT_HEADER_ALIASES
        and normalized[1] in REFDES_HEADER_ALIASES
        and normalized[2] in STATUS_HEADER_ALIASES
    ):
        return "status3"
    if len(normalized) == 4 and (
        normalized[0] in COMPONENT_HEADER_ALIASES
        and normalized[1] in REFDES_HEADER_ALIASES
        and normalized[2] in STATUS_HEADER_ALIASES
        and normalized[3] in NET_HEADER_ALIASES
    ):
        return "status4"
    return None


def _looks_like_refdes_status_rows(rows: list[tuple[int, list[str]]]) -> bool:
    if not rows:
        return False
    values = rows[0][1]
    nonempty_values = [value.strip() for value in values if value.strip()]
    if _refdes_header_kind(values) in {"status2", "status3", "status4"}:
        return True
    if len(nonempty_values) >= 3:
        return True
    if len(nonempty_values) != 2:
        return False
    # Headerless RefDes/status files are status imports when at least one
    # second-column value is status-like (case-insensitive). This deliberately
    # routes mixed valid/invalid status rows to the status parser so they are
    # rejected atomically instead of being treated as component changes.
    status_like_values = [
        row_values[1].strip()
        for _, row_values in rows
        if len(row_values) > 1 and row_values[1].strip()
    ]
    return any(value.casefold() in IMPORTABLE_ACTIVATION_STATUS_TOKENS for value in status_like_values)


def _is_ambiguous_refdes_status_rows(
    rows: list[tuple[int, list[str]]],
    refdes_records: list[RefDesRecord],
    blocks: list[PartialCktBlock],
) -> bool:
    if not rows or _refdes_header_kind(rows[0][1]) is not None:
        return False
    first_values = rows[0][1]
    first_trimmed = list(first_values)
    while first_trimmed and not first_trimmed[-1].strip():
        first_trimmed.pop()
    if len(first_trimmed) != 2:
        return False
    known_components = {block.component_name.casefold() for block in blocks}
    known_components.update(record.component_name.casefold() for record in refdes_records)
    status_like_values = [
        row_values[1].strip()
        for _, row_values in rows
        if len(row_values) > 1 and row_values[1].strip()
    ]
    return any(
        value.casefold() in IMPORTABLE_ACTIVATION_STATUS_TOKENS and value.casefold() in known_components
        for value in status_like_values
    )


def _format_bytes(value: int) -> str:
    size = float(value)
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} B"
        size /= 1024
    return f"{size:.1f} GB"
