from pathlib import Path

from openpyxl import load_workbook

from spd_model_injector.core.refdes_export import export_refdes_xlsx
from spd_model_injector.core.spd import RefDesRecord


def test_export_refdes_xlsx_writes_headers_and_records_in_order(tmp_path: Path) -> None:
    output_path = tmp_path / "refdes.xlsx"
    records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic", net_name="5V_A"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled", net_name="GND"),
        RefDesRecord(component_name="RES_0402", refdes_name="R1", activation_status="Disabled", net_name=""),
    ]

    export_refdes_xlsx(output_path, records)

    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows == [
        ("Component", "RefDes Name", "Activation Status", "Net Name"),
        ("CAP_0402", "C100_0", "Automatic", "5V_A"),
        ("CAP_0402", "C285_0", "Enabled", "GND"),
        ("RES_0402", "R1", "Disabled", None),
    ]
