import os
import time
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QEventLoop, QItemSelectionModel, QTimer, Qt, QPoint
from PySide6.QtWidgets import QApplication, QAbstractItemView, QFrame, QHeaderView, QMessageBox, QPlainTextEdit, QSplitter, QTabWidget

from spd_model_injector import __version__
from spd_model_injector.core.spd import PartialCktBlock, PortRecord, RefDesRecord, SpdInventory
from spd_model_injector.core.spd import PortRequest
from spd_model_injector.ui.main_window import MainWindow
from spd_model_injector.ui.workers import ExportWorker


def _spin_until(app: QApplication, predicate, timeout: float, what: str) -> None:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        app.processEvents()
        if predicate():
            return
        time.sleep(0.005)
    raise AssertionError(f"Timed out after {timeout}s waiting for: {what}")


def _make_block(name: str, ext_nodes: list[str] | None = None) -> PartialCktBlock:
    ext_nodes = ext_nodes if ext_nodes is not None else ["1", "2"]
    return PartialCktBlock(
        component_name=name,
        ext_nodes=ext_nodes,
        start_line=1,
        end_line=3,
        block_start_offset=0,
        body_start_offset=30,
        body_end_offset=40,
        block_end_offset=55,
        header_lines=[f".PartialCkt {name} ExtNode =  " + " ".join(ext_nodes)],
    )


def test_main_window_has_expected_title_and_empty_initial_state() -> None:
    app = QApplication.instance() or QApplication([])

    window = MainWindow()

    assert app is not None
    assert window.windowTitle() == f"SPD Model Injector {__version__}"
    assert window.component_list.count() == 0
    assert "Load an SPD file" in window.status_label.text()
    assert window.undo_component_change_action is not None
    assert window.undo_component_change_action.shortcut().toString() == "Ctrl+Z"


def test_main_window_places_refdes_list_in_right_side_work_area() -> None:
    app = QApplication.instance() or QApplication([])

    window = MainWindow()
    root = window.centralWidget()
    work_splitter = window.findChild(QSplitter, "work_splitter")
    port_splitter = window.findChild(QSplitter, "port_splitter")

    tabs = window.workspace_tabs
    assert root is not tabs
    assert tabs.count() == 2
    assert tabs.tabPosition() == QTabWidget.TabPosition.East
    assert [tabs.tabText(i) for i in range(tabs.count())] == ["Model & RefDes", "Port Generation"]
    assert work_splitter is not None
    assert work_splitter.orientation() == Qt.Orientation.Horizontal
    assert work_splitter.count() == 2
    assert port_splitter is not None
    assert port_splitter.orientation() == Qt.Orientation.Horizontal
    assert port_splitter.count() == 2
    assert port_splitter.sizes()[0] >= port_splitter.sizes()[1]
    assert work_splitter.widget(0).findChild(type(window.editor)) is window.editor
    assert work_splitter.widget(1).findChild(type(window.refdes_table)) is window.refdes_table
    assert window.status_log.parent() is not work_splitter.widget(0)
    assert window.status_log.parent() is not work_splitter.widget(1)
    menu_actions = window.menuBar().actions()
    assert [action.text() for action in menu_actions] == ["File", "Edit", "Model", "Port", "View", "Help"]
    assert [action.text() for action in menu_actions[0].menu().actions()] == [
        "Load SPD", "Export New SPD", "", "Export RefDes Excel", "Import RefDes Excel", "", "Exit"
    ]
    assert [action.text() for action in menu_actions[2].menu().actions()] == ["Validate/Convert", "Revert"]
    assert [action.text() for action in menu_actions[3].menu().actions()] == ["Generate Port", "Clear Pending Ports"]
    window.resize(900, 600)
    window.show()
    app.processEvents()
    window.port_workspace_action.trigger()
    assert window.workspace_tabs.currentIndex() == 1
    assert window.port_workspace_action.isChecked() and not window.model_workspace_action.isChecked()
    assert window.status_log.isVisible()
    window.port_workspace_action.trigger()
    assert window.port_workspace_action.isChecked()
    window.model_workspace_action.trigger()
    assert window.workspace_tabs.currentIndex() == 0
    assert window.model_workspace_action.isChecked() and not window.port_workspace_action.isChecked()


def test_refdes_menu_and_drop_share_auto_detection_and_help_action(monkeypatch, tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.refdes_records = [RefDesRecord(component_name="CAP_0402", refdes_name="C1", activation_status="Enabled")]
    selected = tmp_path / "input.xlsx"
    selected.write_bytes(b"")
    calls: list[Path] = []
    monkeypatch.setattr(window, "import_refdes_file", lambda path: calls.append(Path(path)))
    monkeypatch.setattr(
        "spd_model_injector.ui.main_window.QFileDialog.getOpenFileName",
        lambda *args: (str(selected), ""),
    )

    assert window.import_refdes_status_action is not None
    window.import_refdes_status_action.setEnabled(True)
    window.import_refdes_status_action.trigger()
    window.import_refdes_drop_file(selected)

    assert calls == [selected, selected]
    help_action = next(action for action in window.help_menu.actions() if action.text() == "Input File Formats")
    messages: list[str] = []
    monkeypatch.setattr(QMessageBox, "information", lambda *args: messages.append(str(args[2])))
    help_action.trigger()
    assert messages and all(token in messages[0] for token in ("Component changes", "Status updates", "optional", "aliases"))
    assert window.import_refdes_status_action is not None
    assert window.import_refdes_status_action.text() == "Import RefDes Excel"


def test_refdes_import_aliases_and_headerless_four_column_status_are_partial(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])

    def make_window() -> MainWindow:
        window = MainWindow()
        window.blocks = [_make_block("CAP_0402"), _make_block("Enabled")]
        window.refdes_records = [
            RefDesRecord(component_name="CAP_0402", refdes_name="C1", activation_status="Automatic"),
            RefDesRecord(component_name="CAP_0402", refdes_name="C2", activation_status="Enabled"),
        ]
        window.rebuild_refdes_groups()
        return window

    def write_xlsx(name: str, rows: list[list[str]]) -> Path:
        path = tmp_path / name
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path

    status_cases = [
        ("two.xlsx", [["  rEfDeS  ", " activation status "], ["C1", "Disabled"]]),
        ("three.xlsx", [["Component Name", " RefDes Name ", "Status"], ["CAP_0402", "C1", "Disabled"]]),
        ("four.xlsx", [["component type", "REFDES", "Status", "NET"], ["CAP_0402", "C1", "Disabled", "ignored"]]),
        ("headerless-four.xlsx", [["CAP_0402", "C1", "Disabled", ""], ["CAP_0402", "C2", "Disabled", "GND"]]),
    ]
    for name, rows in status_cases:
        window = make_window()
        path = write_xlsx(name, rows)
        assert window.validate_refdes_status_file(path) == []
        window.import_refdes_file(path)
        assert window.effective_activation_status_for_refdes("C1") == "Disabled"
        expected_c2 = "Disabled" if name == "headerless-four.xlsx" else "Enabled"
        assert window.effective_activation_status_for_refdes("C2") == expected_c2

    component_window = make_window()
    component_path = write_xlsx("component-alias.xlsx", [[" RefDes ", " Component Name "], ["C1", "Enabled"]])
    component_window.import_refdes_file(component_path)
    assert component_window.refdes_component_changes == {"C1": "Enabled"}


def test_main_window_refdes_list_has_framed_sortable_resizable_columns() -> None:
    QApplication.instance() or QApplication([])

    window = MainWindow()
    header = window.refdes_table.horizontalHeader()

    assert window.refdes_panel.frameShape() != QFrame.Shape.NoFrame
    assert window.refdes_panel.layout().indexOf(window.refdes_table) >= 0
    assert window.refdes_table.isSortingEnabled()
    assert header.sectionsClickable()
    assert header.isSortIndicatorShown()
    assert header.sectionResizeMode(0) == QHeaderView.ResizeMode.Interactive
    assert header.sectionResizeMode(1) == QHeaderView.ResizeMode.Interactive
    assert window.refdes_table.selectionMode() == QAbstractItemView.SelectionMode.ExtendedSelection
    assert window.refdes_table.acceptDrops()


def test_main_window_import_model_text_maps_selected_block_ports(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    block = _make_block("C1")
    window.spd_path = tmp_path / "board.spd"
    window.blocks = [block]
    window.populate_components()
    assert window.component_list.item(0).text().startswith("C1")
    window.component_list.setCurrentRow(0)

    window.import_model_text(".SUBCKT CAP Port1 Port2\nC1 Port1 Port2 1u\n.ENDS CAP\n")

    assert window.editor.toPlainText() == "C1 1 2 1u\n"
    assert window.replacements == {"C1": "C1 1 2 1u\n"}
    assert "OK" in window.validation_label.text()


def test_main_window_refdes_table_follows_selected_component(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    cap_block = _make_block("CAP_0402")
    res_block = PartialCktBlock(
        component_name="RES_0402",
        ext_nodes=["A", "B"],
        start_line=4,
        end_line=6,
        block_start_offset=56,
        body_start_offset=90,
        body_end_offset=100,
        block_end_offset=115,
        header_lines=[".PartialCkt RES_0402 ExtNode =  A B"],
    )
    window.spd_path = tmp_path / "board.spd"
    window.inventory = SpdInventory(
        blocks=[cap_block, res_block],
        refdes_records=[
            RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
            RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
            RefDesRecord(component_name="RES_0402", refdes_name="R1", activation_status="Disabled"),
        ],
    )
    window.blocks = window.inventory.blocks
    window.refdes_records = window.inventory.refdes_records
    window.refdes_by_component = window.inventory.refdes_by_component
    window.populate_components()

    window.component_list.setCurrentRow(0)

    assert window.refdes_table.rowCount() == 2
    assert window.refdes_table.item(0, 0).text() == "C100_0"
    assert window.refdes_table.item(0, 1).text() == "Automatic"
    assert "CAP_0402" in window.refdes_label.text()

    window.component_list.setCurrentRow(1)

    assert window.refdes_table.rowCount() == 1
    assert window.refdes_table.item(0, 0).text() == "R1"
    assert window.refdes_table.item(0, 1).text() == "Disabled"
    assert "RES_0402" in window.refdes_label.text()


def test_main_window_changes_multiple_refdes_components_and_undoes_batch(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    cap_block = _make_block("CAP_0402")
    res_block = PartialCktBlock(
        component_name="RES_0402",
        ext_nodes=["A", "B"],
        start_line=4,
        end_line=6,
        block_start_offset=56,
        body_start_offset=90,
        body_end_offset=100,
        block_end_offset=115,
        header_lines=[".PartialCkt RES_0402 ExtNode =  A B"],
    )
    window.spd_path = tmp_path / "board.spd"
    window.inventory = SpdInventory(
        blocks=[cap_block, res_block],
        refdes_records=[
            RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
            RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
            RefDesRecord(component_name="RES_0402", refdes_name="R1", activation_status="Disabled"),
        ],
    )
    window.blocks = window.inventory.blocks
    window.refdes_records = window.inventory.refdes_records
    window.refdes_by_component = window.inventory.refdes_by_component
    window.populate_components()
    window.component_list.setCurrentRow(0)

    first = window.refdes_table.model().index(0, 0)
    second = window.refdes_table.model().index(1, 0)
    selection = window.refdes_table.selectionModel()
    selection.select(first, QItemSelectionModel.SelectionFlag.Select | QItemSelectionModel.SelectionFlag.Rows)
    selection.select(second, QItemSelectionModel.SelectionFlag.Select | QItemSelectionModel.SelectionFlag.Rows)
    window.apply_refdes_component_changes(window.selected_refdes_names(), "RES_0402")

    assert window.refdes_component_changes == {"C100_0": "RES_0402", "C285_0": "RES_0402"}
    assert window.refdes_table.rowCount() == 0
    assert window.undo_component_change_action is not None
    assert window.undo_component_change_action.isEnabled()

    window.undo_refdes_component_change()

    assert window.refdes_component_changes == {}
    assert window.refdes_table.rowCount() == 2
    assert not window.undo_component_change_action.isEnabled()


def test_main_window_refdes_excel_export_uses_changed_components(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [
        PartialCktBlock("CAP_0402", ["1"], 1, 2, 0, 10, 11, 12, [".PartialCkt CAP_0402 ExtNode = 1"]),
        PartialCktBlock("CAP_0603", ["1"], 3, 4, 13, 20, 21, 22, [".PartialCkt CAP_0603 ExtNode = 1"]),
    ]
    window.refdes_records = [RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic")]
    window.rebuild_refdes_groups()
    window.apply_refdes_component_changes(["C100_0"], "CAP_0603")
    output_path = tmp_path / "refdes.xlsx"

    window.export_refdes_excel(output_path)

    workbook = load_workbook(output_path)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()

    assert rows == [
        ("Component", "RefDes Name", "Activation Status", "Net Name"),
        ("CAP_0603", "C100_0", "Automatic", None),
    ]


def test_main_window_imports_refdes_status_excel_with_legacy_header(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402"), _make_block("RES_0402")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
        RefDesRecord(component_name="RES_0402", refdes_name="R1", activation_status="Disabled"),
    ]
    window.rebuild_refdes_groups()
    window.populate_components()
    window.component_list.setCurrentRow(0)
    status_path = tmp_path / "status.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component", "RefDes Name", "Activation Status"])
    sheet.append(["CAP_0402", "C100_0", "Enabled"])
    sheet.append(["CAP_0402", "C285_0", "Disabled"])
    sheet.append(["RES_0402", "R1", "Automatic"])
    workbook.save(status_path)

    window.import_refdes_status_file(status_path)

    assert window.refdes_activation_status_changes == {
        "C100_0": "Enabled",
        "C285_0": "Disabled",
        "R1": "Automatic",
    }
    assert [(record.refdes_name, record.activation_status) for record in window.effective_refdes_records()] == [
        ("C100_0", "Enabled"),
        ("C285_0", "Disabled"),
        ("R1", "Automatic"),
    ]
    assert window.refdes_table.item(0, 1).text() == "Enabled"
    assert window.refdes_table.item(1, 1).text() == "Disabled"
    assert "Imported 3 RefDes activation status" in window.status_log.toPlainText()


def test_main_window_imports_four_column_refdes_export(tmp_path: Path, monkeypatch) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic", net_name="5V_A"),
    ]
    status_path = tmp_path / "status_export.xlsx"
    window.export_refdes_excel(status_path)

    # A user should be able to edit the exported workbook in place and import
    # it again.  Keep the nonblank Net Name column from the export so this
    # exercises the real four-column round trip rather than a hand-authored
    # three-column status file.
    workbook = load_workbook(status_path)
    sheet = workbook.active
    assert tuple(cell.value for cell in sheet[1]) == (
        "Component",
        "RefDes Name",
        "Activation Status",
        "Net Name",
    )
    assert sheet.cell(row=2, column=4).value == "5V_A"
    sheet.cell(row=2, column=3).value = "Disabled"
    workbook.save(status_path)
    workbook.close()

    warnings: list[str] = []
    monkeypatch.setattr(QMessageBox, "warning", lambda _parent, _title, message: warnings.append(message))

    # Use the same auto-routing path used by the RefDes table drag/drop UI.
    window.import_refdes_drop_file(status_path)

    assert window.refdes_activation_status_changes == {"C100_0": "Disabled"}
    assert window.effective_activation_status_for_refdes("C100_0") == "Disabled"
    assert window.validate_refdes_status_file(status_path) == []
    assert warnings == []


def test_main_window_accepts_partial_refdes_status_excel_and_rejects_unexpected_names(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.rebuild_refdes_groups()
    mismatch_path = tmp_path / "status_mismatch.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component", "RefDes Name", "Activation Status"])
    sheet.append(["CAP_0402", "C100_0", "Disabled"])
    sheet.append(["CAP_0402", "C999_0", "Enabled"])
    workbook.save(mismatch_path)
    partial_path = tmp_path / "status_partial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component type", "REFDES", "Status"])
    sheet.append(["CAP_0402", "C100_0", "Disabled"])
    workbook.save(partial_path)
    legacy_partial_path = tmp_path / "status_legacy_partial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component", "RefDes Name", "Activation Status"])
    sheet.append(["CAP_0402", "C100_0", "Disabled"])
    workbook.save(legacy_partial_path)

    assert "Unexpected RefDes: C999_0" in window.validate_refdes_status_file(mismatch_path)
    assert window.validate_refdes_status_file(partial_path) == []
    assert window.validate_refdes_status_file(legacy_partial_path) == []

    window.import_refdes_status_file(partial_path)

    assert window.refdes_activation_status_changes == {"C100_0": "Disabled"}
    assert window.effective_activation_status_for_refdes("C285_0") == "Enabled"


def test_main_window_imports_c01_style_refdes_status_excel_as_partial_update(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402_100NF")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402_100NF", refdes_name="C1077_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402_100NF", refdes_name="C164_0", activation_status="Disabled"),
        RefDesRecord(component_name="CAP_0402_100NF", refdes_name="C1149_0", activation_status="Enabled"),
        RefDesRecord(component_name="CAP_0402_100NF", refdes_name="C1094_0", activation_status="Automatic"),
    ]
    window.rebuild_refdes_groups()
    status_path = tmp_path / "c01_style.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component type", "REFDES", "Status"])
    sheet.append(["CAP_0402_100NF", "C1077_0", "Enabled"])
    sheet.append(["CAP_0402_100NF", "C164_0", "Automatic"])
    sheet.append(["CAP_0402_100NF", "C1149_0", "Disabled"])
    workbook.save(status_path)

    window.import_refdes_status_file(status_path)

    assert window.refdes_activation_status_changes == {
        "C1077_0": "Enabled",
        "C164_0": "Automatic",
        "C1149_0": "Disabled",
    }
    assert window.effective_activation_status_for_refdes("C1094_0") == "Automatic"


def test_main_window_refdes_status_import_keeps_strict_row_validation(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]

    def write_status_file(name: str, rows: list[list[str]]) -> Path:
        path = tmp_path / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Component type", "REFDES", "Status"])
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path

    duplicate_path = write_status_file(
        "status_duplicate.xlsx",
        [["CAP_0402", "C100_0", "Enabled"], ["CAP_0402", "C100_0", "Disabled"]],
    )
    blank_path = write_status_file("status_blank.xlsx", [["CAP_0402", "", "Disabled"]])
    extra_path = write_status_file("status_extra.xlsx", [["CAP_0402", "C100_0", "Disabled", "extra"]])
    wrong_case_path = write_status_file("status_wrong_case.xlsx", [["CAP_0402", "C100_0", "disabled"]])
    header_only_path = write_status_file("status_header_only.xlsx", [])

    assert "Duplicate RefDes: C100_0" in window.validate_refdes_status_file(duplicate_path)
    assert "Row 2: Component, RefDes Name, and Activation Status are required." in window.validate_refdes_status_file(blank_path)
    assert "Row 2: expected exactly 3 columns." in window.validate_refdes_status_file(extra_path)
    assert "Unknown Activation Status: disabled" in window.validate_refdes_status_file(wrong_case_path)
    assert window.validate_refdes_status_file(header_only_path) == ["RefDes status file has no data rows."]
    assert window.refdes_activation_status_changes == {}


def test_main_window_refdes_status_import_rejects_component_mismatch_atomically(
    tmp_path: Path,
    monkeypatch,
) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.refdes_activation_status_changes = {"C285_0": "Disabled"}
    status_path = tmp_path / "status_component_mismatch.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component type", "REFDES", "Status"])
    sheet.append(["CAP_0402", "C100_0", "Disabled"])
    sheet.append(["WRONG_COMPONENT", "C285_0", "Automatic"])
    workbook.save(status_path)
    warnings: list[str] = []
    monkeypatch.setattr(QMessageBox, "warning", lambda _parent, _title, message: warnings.append(message))

    window.import_refdes_status_file(status_path)

    assert window.refdes_activation_status_changes == {"C285_0": "Disabled"}
    assert warnings == ["Component mismatch for C285_0: expected CAP_0402, got WRONG_COMPONENT"]


def test_main_window_refdes_drop_routes_status_and_component_xlsx(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402"), _make_block("CAP_0603")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.rebuild_refdes_groups()

    status_path = tmp_path / "status_drop.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component type", "REFDES", "Status"])
    sheet.append(["CAP_0402", "C100_0", "Disabled"])
    workbook.save(status_path)
    component_path = tmp_path / "component_drop.xlsx"
    workbook = Workbook()
    workbook.active.append(["C285_0", "CAP_0603"])
    workbook.save(component_path)

    window.import_refdes_drop_file(status_path)
    window.import_refdes_drop_file(component_path)

    assert window.refdes_activation_status_changes == {"C100_0": "Disabled"}
    assert window.refdes_component_changes == {"C285_0": "CAP_0603"}


def test_main_window_accepts_casefold_status_headers_and_explicit_two_column_header(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    cases = [
        (("Component", "RefDes Name", "Activation Status"), 3, False),
        (("Component type", "REFDES", "Status"), 3, False),
        (("RefDes", "Status"), 2, False),
        (("Component", "RefDes Name", "Activation Status", "Net Name"), 4, False),
    ]
    for index, (header, column_count, requires_full) in enumerate(cases):
        window = MainWindow()
        window.blocks = [_make_block("CAP_0402")]
        window.refdes_records = [
            RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
            RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
        ]
        window.rebuild_refdes_groups()
        path = tmp_path / f"casefold_{index}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([f"  {value.lower()}  " for value in header])
        if column_count == 4:
            sheet.append(["CAP_0402", "C100_0", "Disabled", "5V_A"])
        elif column_count == 3:
            sheet.append(["CAP_0402", "C100_0", "Disabled"])
        else:
            sheet.append(["C100_0", "Disabled"])
        if column_count == 4:
            sheet.append(["CAP_0402", "C285_0", "Automatic", "GND"])
        elif requires_full:
            sheet.append(["CAP_0402", "C285_0", "Automatic"])
        workbook.save(path)

        statuses, _, parsed_requires_full, parse_errors = window._parse_refdes_status_file(path)
        assert statuses and parse_errors == []
        assert parsed_requires_full is requires_full
        assert window.validate_refdes_status_file(path) == []
        window.import_refdes_status_file(path)
        assert window.effective_activation_status_for_refdes("C100_0") == "Disabled"
        assert window.refdes_component_changes == {}


def test_main_window_refdes_drop_applies_headerless_two_and_three_column_status_files(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402"), _make_block("CAP_0603")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.rebuild_refdes_groups()
    three_col = tmp_path / "headerless_three.xlsx"
    workbook = Workbook()
    workbook.active.append(["CAP_0402", "C100_0", "Disabled"])
    workbook.save(three_col)
    two_col = tmp_path / "headerless_two.xlsx"
    workbook = Workbook()
    workbook.active.append(["C285_0", "Disabled"])
    workbook.save(two_col)

    window.import_refdes_drop_file(three_col)
    window.import_refdes_drop_file(two_col)

    assert window.refdes_activation_status_changes == {"C100_0": "Disabled", "C285_0": "Disabled"}
    assert window.refdes_component_changes == {}


def test_main_window_refdes_drop_status_like_two_column_inputs_reject_atomically(tmp_path: Path, monkeypatch) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402"), _make_block("CAP_0603")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.rebuild_refdes_groups()
    warnings: list[str] = []
    monkeypatch.setattr(QMessageBox, "warning", lambda _parent, _title, message: warnings.append(message))

    lowercase = tmp_path / "lowercase_status.xlsx"
    workbook = Workbook()
    workbook.active.append(["C100_0", "disabled"])
    workbook.save(lowercase)
    mixed = tmp_path / "mixed_status.xlsx"
    workbook = Workbook()
    workbook.active.append(["C100_0", "Enabled"])
    workbook.active.append(["C285_0", "CAP_0603"])
    workbook.save(mixed)

    window.import_refdes_drop_file(lowercase)
    window.import_refdes_drop_file(mixed)

    assert len(warnings) == 2
    assert all("Unknown Activation Status" in message for message in warnings)
    assert window.refdes_activation_status_changes == {}
    assert window.refdes_component_changes == {}

    unknown = tmp_path / "unknown_refdes_two.xlsx"
    workbook = Workbook()
    workbook.active.append(["C999_0", "Enabled"])
    workbook.save(unknown)
    assert any("Unexpected RefDes: C999_0" in error for error in window.validate_refdes_status_file(unknown))

    duplicate = tmp_path / "duplicate_refdes_two.xlsx"
    workbook = Workbook()
    workbook.active.append(["C100_0", "Enabled"])
    workbook.active.append(["C100_0", "Disabled"])
    workbook.save(duplicate)
    assert "Duplicate RefDes: C100_0" in window.validate_refdes_status_file(duplicate)

    extra = tmp_path / "extra_refdes_two.xlsx"
    workbook = Workbook()
    workbook.active.append(["C100_0", "Enabled"])
    workbook.active.append(["C285_0", "Disabled", "extra"])
    workbook.save(extra)
    assert "Row 2: expected exactly 2 columns." in window.validate_refdes_status_file(extra)


def test_main_window_refdes_drop_rejects_two_column_status_component_collision(tmp_path: Path, monkeypatch) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("Enabled")]
    window.refdes_records = [RefDesRecord(component_name="Enabled", refdes_name="C100_0", activation_status="Automatic")]
    window.rebuild_refdes_groups()
    path = tmp_path / "ambiguous_two.xlsx"
    workbook = Workbook()
    workbook.active.append(["C100_0", "Enabled"])
    workbook.save(path)
    warnings: list[str] = []
    monkeypatch.setattr(QMessageBox, "warning", lambda _parent, _title, message: warnings.append(message))

    window.import_refdes_drop_file(path)

    assert len(warnings) == 1
    assert "Ambiguous 2-column" in warnings[0]
    assert window.refdes_activation_status_changes == {}
    assert window.refdes_component_changes == {}


def test_main_window_refdes_drop_rejects_malformed_status_header_once(tmp_path: Path, monkeypatch) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic")
    ]
    status_path = tmp_path / "status_bad_header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component type", "REFDES", "State"])
    sheet.append(["CAP_0402", "C100_0", "Disabled"])
    workbook.save(status_path)
    warnings: list[str] = []
    monkeypatch.setattr(QMessageBox, "warning", lambda _parent, _title, message: warnings.append(message))

    window.import_refdes_drop_file(status_path)

    assert len(warnings) == 1
    assert warnings[0].startswith("Row 1: expected header:")
    assert window.refdes_activation_status_changes == {}
    assert window.refdes_component_changes == {}


def test_main_window_refdes_status_import_round_trips_unknown_only_for_unknown_records(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Unknown"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.rebuild_refdes_groups()
    status_path = tmp_path / "status_unknown.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component", "RefDes Name", "Activation Status"])
    sheet.append(["CAP_0402", "C100_0", "Unknown"])
    sheet.append(["CAP_0402", "C285_0", "Disabled"])
    workbook.save(status_path)
    invalid_path = tmp_path / "status_invalid_unknown.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Component", "RefDes Name", "Activation Status"])
    sheet.append(["CAP_0402", "C100_0", "Enabled"])
    sheet.append(["CAP_0402", "C285_0", "Unknown"])
    workbook.save(invalid_path)

    window.import_refdes_status_file(status_path)

    assert window.refdes_activation_status_changes == {"C285_0": "Disabled"}
    assert "Unknown Activation Status: Unknown" in window.validate_refdes_status_file(invalid_path)


def test_main_window_refdes_status_import_rejects_unreadable_xlsx(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.refdes_records = [RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic")]
    bad_path = tmp_path / "not_excel.xlsx"
    bad_path.write_text("not an excel workbook", encoding="utf-8")

    errors = window.validate_refdes_status_file(bad_path)

    assert len(errors) == 1
    assert "Could not read RefDes status file" in errors[0]


def test_main_window_changes_selected_refdes_activation_statuses(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("CAP_0402")]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.rebuild_refdes_groups()
    window.populate_components()
    window.component_list.setCurrentRow(0)

    first = window.refdes_table.model().index(0, 0)
    second = window.refdes_table.model().index(1, 0)
    selection = window.refdes_table.selectionModel()
    selection.select(first, QItemSelectionModel.SelectionFlag.Select | QItemSelectionModel.SelectionFlag.Rows)
    selection.select(second, QItemSelectionModel.SelectionFlag.Select | QItemSelectionModel.SelectionFlag.Rows)
    window.apply_refdes_activation_status_changes(window.selected_refdes_names(), "Disabled")

    assert window.refdes_activation_status_changes == {"C100_0": "Disabled", "C285_0": "Disabled"}
    assert window.refdes_table.item(0, 1).text() == "Disabled"
    assert window.refdes_table.item(1, 1).text() == "Disabled"


def test_export_worker_preserves_refdes_record_fallback_scan_for_status_changes(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    spd_path = tmp_path / "board.spd"
    output_path = tmp_path / "board_out.spd"
    spd_path.write_text(
        ".Connect C100_0 CAP_0402 Checked = 1\n",
        encoding="utf-8",
        newline="\n",
    )
    finished: list[str] = []
    failed: list[str] = []
    worker = ExportWorker(
        spd_path,
        output_path,
        [],
        {},
        refdes_activation_status_changes={"C100_0": "Enabled"},
    )
    worker.finished.connect(finished.append)
    worker.failed.connect(failed.append)

    worker.run()

    assert failed == []
    assert finished == [str(output_path)]
    assert output_path.read_text(encoding="utf-8") == ".Connect C100_0 CAP_0402 Usage = 0b1000 Checked = 1\n"


def test_main_window_refdes_component_undo_stack_keeps_ten_batches(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    block_a = PartialCktBlock("A", ["1"], 1, 2, 0, 10, 11, 12, [".PartialCkt A ExtNode =  1"])
    block_b = PartialCktBlock("B", ["1"], 3, 4, 13, 20, 21, 22, [".PartialCkt B ExtNode =  1"])
    window.blocks = [block_a, block_b]
    window.refdes_records = [RefDesRecord(component_name="A", refdes_name=f"C{i}", activation_status="Automatic") for i in range(11)]
    window.rebuild_refdes_groups()

    for i in range(11):
        window.apply_refdes_component_changes([f"C{i}"], "B")

    assert len(window.refdes_component_undo_stack) == 10
    assert window.refdes_component_undo_stack[0].changes[0].refdes_name == "C1"


def test_main_window_imports_refdes_component_csv_and_xlsx_without_headers(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [
        PartialCktBlock("CAP_0402", ["1"], 1, 2, 0, 10, 11, 12, [".PartialCkt CAP_0402 ExtNode = 1"]),
        PartialCktBlock("CAP_0603", ["1"], 3, 4, 13, 20, 21, 22, [".PartialCkt CAP_0603 ExtNode = 1"]),
    ]
    window.refdes_records = [
        RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic"),
        RefDesRecord(component_name="CAP_0402", refdes_name="C285_0", activation_status="Enabled"),
    ]
    window.rebuild_refdes_groups()
    csv_path = tmp_path / "changes.csv"
    csv_path.write_text("C100_0,CAP_0603\n", encoding="utf-8")
    xlsx_path = tmp_path / "changes.xlsx"
    workbook = Workbook()
    workbook.active.append(["C285_0", "CAP_0603"])
    workbook.save(xlsx_path)

    assert window.load_refdes_component_change_file(csv_path) == {"C100_0": "CAP_0603"}
    assert window.load_refdes_component_change_file(xlsx_path) == {"C285_0": "CAP_0603"}


def test_main_window_rejects_invalid_refdes_component_import_without_changes(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [PartialCktBlock("CAP_0402", ["1"], 1, 2, 0, 10, 11, 12, [".PartialCkt CAP_0402 ExtNode = 1"])]
    window.refdes_records = [RefDesRecord(component_name="CAP_0402", refdes_name="C100_0", activation_status="Automatic")]
    window.rebuild_refdes_groups()
    duplicate_path = tmp_path / "duplicate.csv"
    duplicate_path.write_text("C100_0,CAP_0402\nC100_0,MISSING\n", encoding="utf-8")
    unknown_component_path = tmp_path / "unknown.csv"
    unknown_component_path.write_text("C100_0,MISSING\n", encoding="utf-8")

    assert window.validate_refdes_component_changes({"C100_0": "MISSING"}) == ["Unknown Component: MISSING"]
    assert "Duplicate RefDes: C100_0" in window.validate_refdes_component_change_file(duplicate_path)
    assert "Unknown Component: MISSING" in window.validate_refdes_component_change_file(unknown_component_path)
    assert window.refdes_component_changes == {}


def test_component_filter_hides_non_matching_rows_and_updates_header() -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    window.blocks = [_make_block("C1"), _make_block("C2"), _make_block("R1")]
    window.populate_components()

    assert window.component_list_label.text() == "PartialCkt Components (3/3)"

    window.component_filter.setText("c")

    visible_rows = [row for row in range(window.component_list.count()) if not window.component_list.item(row).isHidden()]
    hidden_rows = [row for row in range(window.component_list.count()) if window.component_list.item(row).isHidden()]
    assert len(visible_rows) == 2
    assert len(hidden_rows) == 1
    assert window.component_list_label.text() == "PartialCkt Components (2/3)"
    assert window.component_list.count() == 3

    window.component_filter.setText("")
    assert window.component_list_label.text() == "PartialCkt Components (3/3)"


def test_modified_item_is_visually_marked_after_import(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    block = _make_block("C1")
    window.spd_path = tmp_path / "board.spd"
    window.blocks = [block]
    window.populate_components()
    window.component_list.setCurrentRow(0)

    item_before = window.component_list.item(0)
    assert not item_before.font().bold()
    assert not item_before.text().startswith("* ")

    window.import_model_text(".SUBCKT CAP Port1 Port2\nC1 Port1 Port2 1u\n.ENDS CAP\n")

    item_after = window.component_list.item(0)
    assert item_after.font().bold()
    assert item_after.text().startswith("* ")
    assert "ports: 2" in item_after.text()


def test_load_block_body_missing_file_shows_error_and_leaves_editor_empty(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    block = _make_block("C1")
    window.spd_path = tmp_path / "missing.spd"
    window.blocks = [block]
    window.populate_components()

    window.component_list.setCurrentRow(0)

    assert window.editor.toPlainText() == ""
    assert "error" in window.validation_label.text().lower() or "could not read" in window.validation_label.text().lower()
    assert "could not read" in window.status_label.text().lower()


def test_busy_state_disables_load_export_validate_actions_and_buttons() -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()

    assert window.load_action.isEnabled()
    assert window.export_action.isEnabled()
    assert window.validate_action.isEnabled()
    assert window.import_button.isEnabled()
    assert window.validate_button.isEnabled()

    window._set_busy(True)

    assert not window.load_action.isEnabled()
    assert not window.export_action.isEnabled()
    assert not window.validate_action.isEnabled()
    assert not window.import_button.isEnabled()
    assert not window.validate_button.isEnabled()

    window._set_busy(False)

    assert window.load_action.isEnabled()
    assert window.export_action.isEnabled()
    assert window.validate_action.isEnabled()
    assert window.import_button.isEnabled()
    assert window.validate_button.isEnabled()


def test_main_window_load_spd_populates_components_from_worker(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    spd_path = tmp_path / "board.spd"
    spd_path.write_text(
        "Title\n"
        ".PartialCkt C1 ExtNode =  1 2\n"
        "C 1 2 1u\n"
        ".EndPartialCkt\n"
        ".Connect C100_0 C1 Checked = 1\n",
        encoding="utf-8",
        newline="\n",
    )
    window = MainWindow()
    loop = QEventLoop()
    original_finished = window._scan_finished

    def finished(inventory: SpdInventory) -> None:
        original_finished(inventory)
        loop.quit()

    window._scan_finished = finished
    window.load_spd(spd_path)
    QTimer.singleShot(3000, loop.quit)
    loop.exec()

    if window._scan_thread is not None and window._scan_thread.isRunning():
        window._scan_thread.quit()
        window._scan_thread.wait(1000)

    assert window.component_list.count() == 1
    assert window.blocks[0].component_name == "C1"
    assert window.refdes_table.rowCount() == 1
    assert window.refdes_table.item(0, 0).text() == "C100_0"
    assert window.refdes_table.item(0, 1).text() == "Automatic"
    assert "Loaded 1 PartialCkt blocks" in window.status_log.toPlainText()


def test_load_spd_real_threaded_scan_completes(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    window = MainWindow()
    spd = tmp_path / "board.spd"
    spd.write_text(
        ".PartialCkt C1 ExtNode =  1 2\nC 1 2 1u\n.EndPartialCkt\n"
        ".PartialCkt U1 ExtNode =  A B\nR A B 1\n.EndPartialCkt\n",
        newline="\n",
    )

    window.load_spd(spd)
    assert window._busy, "busy state should be set while the scan runs"
    assert window._scan_worker is not None, "worker must be strongly referenced during the scan"

    _spin_until(app, lambda: not window._busy, timeout=15.0, what="threaded scan to finish")

    assert window.spd_path == spd
    assert len(window.blocks) == 2
    assert window.component_list.count() == 2
    assert "2/2" in window.component_list_label.text()
    assert window.load_action.isEnabled()

    _spin_until(
        app,
        lambda: window._scan_thread is None and window._scan_worker is None,
        timeout=15.0,
        what="scan thread/worker refs to be cleared",
    )


def test_editor_uses_fixed_pitch_font_and_no_wrap() -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()

    assert window.editor.font().fixedPitch()
    assert window.editor.lineWrapMode() == QPlainTextEdit.LineWrapMode.NoWrap


def test_generate_port_queues_selected_refdes_and_export_worker_forwards_requests(tmp_path: Path, monkeypatch) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    source = tmp_path / "board.spd"
    source.write_text("source\n", encoding="utf-8")
    record = RefDesRecord(
        component_name="CAP",
        refdes_name="C1",
        activation_status="Automatic",
        net_name="VDD",
        unique_net_names=("DGND", "VDD"),
        package_node_count=2,
        annotated_node_count=2,
    )
    record2 = RefDesRecord("CAP", "C2", "Automatic", net_name="VDD", unique_net_names=("DGND", "VDD"), package_node_count=2, annotated_node_count=2)
    window.spd_path = source
    window.inventory = SpdInventory(
        [], [record, record2], ground_nets=("DGND",), net_names=("DGND", "VDD"), existing_port_keys=(), port_insertion_offset=1,
    )
    window.blocks = [_make_block("CAP")]
    window.refdes_records = [record, record2]
    window.rebuild_refdes_groups()
    window._set_net_selectors(("VDD",), ("DGND", "VDD"), ("DGND",))
    window.power_net_list.item(0).setCheckState(Qt.CheckState.Checked)
    window._populate_refdes_table("CAP")
    window.refdes_table.selectRow(0)

    assert window.port_refdes_table.topLevelItemCount() == 1
    assert window.port_refdes_table.topLevelItem(0).childCount() == 2
    window.port_refdes_table.expandAll()
    assert window.selected_refdes_names() == ["C1"]
    assert not window._selected_port_refdes_names()
    assert not window.generate_port_action.isEnabled()
    assert not window.port_export_button.isEnabled()
    first = window.port_refdes_table.topLevelItem(0).child(0)
    second = window.port_refdes_table.topLevelItem(0).child(1)
    window.port_refdes_table.setCurrentItem(first)
    window.port_refdes_table.selectionModel().select(
        window.port_refdes_table.indexFromItem(second, 0), QItemSelectionModel.SelectionFlag.Select | QItemSelectionModel.SelectionFlag.Rows
    )

    assert window.generate_port_action.isEnabled()
    window._set_busy(True)
    assert not window.generate_port_action.isEnabled()
    window._set_busy(False)
    window.inventory = SpdInventory([], [record, record2], ground_nets=("DGND",), net_names=("DGND", "VDD"))
    window._update_generate_port_state()
    assert not window.generate_port_action.isEnabled()
    window.inventory = SpdInventory([], [record, record2], ground_nets=("DGND",), net_names=("DGND", "VDD"), port_insertion_offset=1)
    window.port_refdes_table.expandAll()
    window.port_refdes_table.selectAll()
    window._update_generate_port_state()
    assert window.generate_port_action.isEnabled()
    monkeypatch.setattr("spd_model_injector.ui.main_window.validate_port_requests", lambda *args, **kwargs: [])
    window.generate_ports()
    assert window.pending_port_requests == [
        PortRequest(instance="C1", target_net="VDD", reference_net="DGND"),
        PortRequest(instance="C2", target_net="VDD", reference_net="DGND"),
    ]
    assert "Queued 2 Port request" in window.status_log.toPlainText()
    assert not window._selected_port_refdes_names()
    assert window.port_clear_button.isEnabled()
    assert window.port_export_button.isEnabled()

    captured: dict[str, object] = {}
    monkeypatch.setattr("spd_model_injector.ui.workers.write_spd_with_replacements", lambda *args, **kwargs: captured.update(kwargs))
    worker = ExportWorker(
        source, tmp_path / "out.spd", window.blocks, {}, port_requests=window.pending_port_requests,
        port_deletions=["Port1_OLD::VDD"], port_enabled_changes={"Port2_OLD::VDD": False}, inventory=window.inventory
    )
    worker.run()
    assert captured["port_requests"] == window.pending_port_requests
    assert captured["port_deletions"] == ["Port1_OLD::VDD"]
    assert captured["port_enabled_changes"] == {"Port2_OLD::VDD": False}
    assert captured["inventory"] is window.inventory

    window.clear_pending_ports()
    assert not window.pending_port_requests
    assert not window.port_clear_button.isEnabled()
    assert not window.port_export_button.isEnabled()
    assert "Cleared 2 pending Port request" in window.status_log.toPlainText()


def test_existing_port_table_queues_activation_delete_and_restore(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    source = tmp_path / "board.spd"
    source.write_text("source\n", encoding="utf-8")
    port = PortRecord(
        name="Port7_U1::VDD", number=7, instance="U1", target_net="VDD", component_name="DUT",
        enabled=False, positive_node_count=12, negative_node_count=8,
        header_start_offset=10, header_end_offset=40, record_end_offset=100,
        header_line="Port7_U1::VDD Disabled Auto\n",
    )
    window.spd_path = source
    window.inventory = SpdInventory([], [], port_records=(port,), port_insertion_offset=100)
    window._populate_port_management_table()

    assert window.port_management_table.rowCount() == 1
    assert window.port_management_table.item(0, 0).checkState() == Qt.CheckState.Unchecked
    assert window.port_management_table.item(0, 4).text() == "12"
    window.port_management_table.item(0, 0).setCheckState(Qt.CheckState.Checked)
    assert window.port_enabled_changes == {port.name: True}
    assert window.port_export_button.isEnabled()

    window.port_management_table.selectRow(0)
    window.delete_or_restore_selected_ports()
    assert window.port_deletions == {port.name}
    assert not window.port_enabled_changes
    assert window.port_management_table.item(0, 6).text() == "Delete queued"
    window.port_management_table.selectRow(0)
    window.delete_or_restore_selected_ports()
    assert not window.port_deletions


def test_port_filter_clears_hidden_selection_and_readiness_distinguishes_busy_and_unsafe(tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    source = tmp_path / "board.spd"
    source.write_text("source\n", encoding="utf-8")
    eligible = RefDesRecord(
        "DUT", "U1", "Automatic", net_name="VDD", unique_net_names=("DGND", "VDD"),
        net_node_counts=(("DGND", 5), ("VDD", 3)), package_node_count=9, annotated_node_count=8,
    )
    ineligible = RefDesRecord(
        "DUT", "U2", "Automatic", net_name="VDD", unique_net_names=("VDD",),
        net_node_counts=(("VDD", 3),), package_node_count=3, annotated_node_count=3,
    )
    window.spd_path = source
    window.refdes_records = [eligible, ineligible]
    window.inventory = SpdInventory([], [eligible, ineligible], net_names=("DGND", "VDD"), port_insertion_offset=1)
    window._set_net_selectors(("VDD",), ("DGND", "VDD"), ("DGND",))
    window.power_net_list.item(0).setCheckState(Qt.CheckState.Checked)

    assert window.port_refdes_table.topLevelItemCount() == 1
    assert window.port_refdes_table.topLevelItem(0).childCount() == 1
    window.port_refdes_table.expandAll()
    assert window.port_candidate_label.text().endswith(": 1")
    leaf = window.port_refdes_table.topLevelItem(0).child(0)
    assert leaf.text(1) == "VDD (3)"
    assert leaf.text(2) == "5"
    window.port_refdes_table.setCurrentItem(leaf)
    assert window.generate_port_action.isEnabled()
    window.port_refdes_filter.setText("does-not-match")
    assert leaf.isHidden()
    assert not window._selected_port_refdes_names()
    assert not window.generate_port_action.isEnabled()

    window._set_busy(True)
    assert window.port_readiness_banner.text().startswith("Busy:")
    window._set_busy(False)
    window.inventory = SpdInventory([], [eligible, ineligible])
    window._update_generate_port_state()
    assert "no safe .Port/.EndPort section" in window.port_readiness_banner.text()


def test_port_workspace_queues_each_checked_power_channel_with_auto_dgnd(monkeypatch, tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    record = RefDesRecord(
        "CAP", "C1", "Automatic", unique_net_names=("DGND", "VDD"),
        package_node_count=2, annotated_node_count=2,
    )
    record2 = RefDesRecord("CAP", "C2", "Automatic", unique_net_names=("DGND", "VTT"), package_node_count=2, annotated_node_count=2)
    window.spd_path = tmp_path / "board.spd"
    window.spd_path.write_text("source\n", encoding="utf-8")
    window.refdes_records = [record, record2]
    window.inventory = SpdInventory(
        [], [record, record2], net_names=("DGND", "VDD", "VTT"), power_nets=("VDD", "VTT"),
        ground_nets=("DGND",), port_insertion_offset=1,
    )
    window.rebuild_refdes_groups()
    window._set_net_selectors(("VDD", "VTT"), ("DGND", "VDD", "VTT"), ("DGND",))
    for row in range(window.power_net_list.count()):
        window.power_net_list.item(row).setCheckState(Qt.CheckState.Checked)
    first = window.port_refdes_table.topLevelItem(0).child(0)
    second = window.port_refdes_table.topLevelItem(0).child(1)
    window.port_refdes_table.setCurrentItem(first)
    window.port_refdes_table.selectionModel().select(
        window.port_refdes_table.indexFromItem(second, 0), QItemSelectionModel.SelectionFlag.Select | QItemSelectionModel.SelectionFlag.Rows
    )
    monkeypatch.setattr("spd_model_injector.ui.main_window.validate_port_requests", lambda *args, **kwargs: [])

    assert window.reference_net_display.text() == "Reference: Auto: DGND"
    window.port_refdes_table.expandAll()
    window.generate_ports()
    assert [(request.instance, request.target_net, request.reference_net) for request in window.pending_port_requests] == [
        ("C1", "VDD", "DGND"), ("C2", "VTT", "DGND")
    ]


def test_port_workspace_prompts_once_for_reference_when_dgnd_missing(monkeypatch, tmp_path: Path) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    record = RefDesRecord("CAP", "C1", "Automatic", unique_net_names=("VDD", "VSS"), package_node_count=2, annotated_node_count=2)
    window.spd_path = tmp_path / "board.spd"
    window.spd_path.write_text("source\n", encoding="utf-8")
    window.refdes_records = [record]
    window.inventory = SpdInventory([], [record], net_names=("VDD", "VSS"), power_nets=("VDD",), ground_nets=("VSS",), port_insertion_offset=1)
    window._set_net_selectors(("VDD",), ("VDD", "VSS"), ("VSS",))
    window.power_net_list.item(0).setCheckState(Qt.CheckState.Checked)
    window.port_refdes_table.expandAll()
    window.port_refdes_table.setCurrentItem(window.port_refdes_table.topLevelItem(0).child(0))
    monkeypatch.setattr("spd_model_injector.ui.main_window.QInputDialog.getItem", lambda *args, **kwargs: ("VSS", True))
    monkeypatch.setattr("spd_model_injector.ui.main_window.validate_port_requests", lambda *args, **kwargs: [])
    window.generate_ports()
    assert window.pending_port_requests[0].reference_net == "VSS"


def test_port_refdes_tree_context_menu_expands_and_collapses(monkeypatch) -> None:
    QApplication.instance() or QApplication([])
    window = MainWindow()
    record = RefDesRecord("CAP", "C1", "Automatic", unique_net_names=("DGND", "VDD"))
    window.refdes_records = [record]
    window.inventory = SpdInventory([], [record], net_names=("DGND", "VDD"), power_nets=("VDD",), port_insertion_offset=1)
    window.spd_path = Path("board.spd")
    window._set_net_selectors(("VDD",), ("DGND", "VDD"), ("DGND",))
    window.power_net_list.item(0).setCheckState(Qt.CheckState.Checked)
    parent = window.port_refdes_table.topLevelItem(0)
    parent.setExpanded(False)

    class FakeAction:
        def __init__(self, text: str) -> None:
            self.text = text

    class FakeMenu:
        selected = "Expand All"

        def __init__(self, _parent) -> None:
            self.actions: list[FakeAction] = []

        def addAction(self, text: str) -> FakeAction:
            action = FakeAction(text)
            self.actions.append(action)
            return action

        def exec(self, _position) -> FakeAction:
            return next(action for action in self.actions if action.text == self.selected)

    monkeypatch.setattr("spd_model_injector.ui.main_window.QMenu", FakeMenu)
    window._show_port_refdes_context_menu(QPoint())
    assert parent.isExpanded()
    FakeMenu.selected = "Collapse All"
    window._show_port_refdes_context_menu(QPoint())
    assert not parent.isExpanded()
